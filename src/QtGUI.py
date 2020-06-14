from PyQt5 import QtWidgets, Qt, QtCore

from src.Doc1 import doc1
from src.Nagryzka02 import Nagryzka02
from src.Zvit3 import Zvit3
from src.convert import ConverterXLS
from src.countC import CountC
from src.design import Ui_MainWindow
from src.pars import pars
from openpyxl import *


class AppGUI(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.pushButton.clicked.connect(self.load_file_rnp)
        self.pushButton_2.clicked.connect(self.load_file)
        self.pushButton_5.clicked.connect(self.compare_files)
        self.pushButton_8.clicked.connect(self.convert_file)
        self.pushButton_6.clicked.connect(self.create_report)
        self.only_file_name = []
        self.only_file_name_rnp = []
        self.default_file_names = []
        self.default_file_names_rnp = []
        self.radioButton.click()
        self.lectors = []
        self.assisstents = []
        self.subjects = []
        self.load_db()

        self.pushButton_plus_asistent.clicked.connect(self.add_assisstant)
        self.pushButton_plus_lector.clicked.connect(self.add_lector)
        self.pushButton_minus_assisstent.clicked.connect(self.minus_assisstant)
        self.pushButton_minus_lector.clicked.connect(self.minus_lector)
        self.pushButton_plus_asistent_db.clicked.connect(self.add_assisstant_db)
        self.pushButton_plus_lector_db.clicked.connect(self.add_lector_db)
        self.pushButton_plus_predmet_db.clicked.connect(self.add_subject_db)
        self.pushButton_minus_assisstent_db.clicked.connect(self.remove_assisstant_db)
        self.pushButton_minus_lector_db.clicked.connect(self.remove_lector_db)
        self.pushButton_minus_predmet_db.clicked.connect(self.remove_subject_db)
        self.pushButton_Save_db.clicked.connect(self.save_db)
        self.pushButton_Calculate_load.clicked.connect(self.calculate_load)
        self.comboBox_Subjects.currentTextChanged.connect(self.on_change_value)
        self.pushButton_Calculate_Load_teachers.clicked.connect(self.calculate_load_teachers)

    def calculate_load_teachers(self):
        try:
            ob_load = Zvit3()
            info = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Information,
                                         "Розподіл навантаження", "Розподілення для фахівців виконано успішно!",
                                         QtWidgets.QMessageBox.Ok)
            info.exec_()
        except:
            error = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning,
                                          "Розподіл навантаження", "Спочатку розподіліть навантаження за всіма формами!",
                                          QtWidgets.QMessageBox.Ok)
            error.exec_()

    def on_change_value(self):
        if self.comboBox.currentText() != '' and self.comboBox_2.currentText() != '' and self.comboBox_4.currentText() != '' and self.comboBox_5.currentText() != '':
            countC_object = CountC(self.default_file_names[self.comboBox.currentIndex()],
                                   self.default_file_names[self.comboBox_4.currentIndex()],
                                   self.default_file_names[self.comboBox_2.currentIndex()],
                                   self.default_file_names[self.comboBox_5.currentIndex()],
                                   self.comboBox_Subjects.currentText())
            _translate = QtCore.QCoreApplication.translate
            self.label_lect_1sem_den.setText(_translate("MainWindow", str(int(countC_object.sem1_den_lektor))))
            self.label_lect_1sem_zaoch.setText(_translate("MainWindow", str(int(countC_object.sem1_zaoch_lektor))))
            self.label_lect_2sem_denna.setText(_translate("MainWindow", str(int(countC_object.sem2_den_lektor))))
            self.label_lect_2sem_zaoch.setText(_translate("MainWindow", str(int(countC_object.sem2_zaoch_lektor))))

            self.label_asist_1sem_den.setText(_translate("MainWindow", str(int(countC_object.sem1_den_ass))))
            self.label_asist_1sem_zaoch.setText(_translate("MainWindow", str(int(countC_object.sem1_zaoch_ass))))
            self.label_asist_2sem_den.setText(_translate("MainWindow", str(int(countC_object.sem2_den_ass))))
            self.label_asist_2sem_zaoch.setText(_translate("MainWindow", str(int(countC_object.sem2_zaoch_ass))))

            self.label_pidgrup_1sem_den.setText(_translate("MainWindow", str(countC_object.podgrup_d1sem)))
            self.label_pidgrup_2sem_den.setText(_translate("MainWindow", str(countC_object.podgrup_d2sem)))
            self.label_pidgrup_1sem_zaoch.setText(_translate("MainWindow", str(countC_object.podgrup_z1sem)))
            self.label_pidgrup_2sem_zaoch.setText(_translate("MainWindow", str(countC_object.podgrup_z2sem)))
        else:
            error = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning,
                                          "Часи навантаження", "Відкрийте файли кафедри!",
                                          QtWidgets.QMessageBox.Ok)
            error.exec_()

    def save_db(self):
        workBookDB = load_workbook("resources\\DB_study_load.xlsx")
        sheet = workBookDB["Sheet1"]
        for i in range(2, sheet.max_row):
            w_cell = sheet.cell(row=i, column=1)
            w_cell.value = ""
        for i in range(2, sheet.max_row):
            w_cell = sheet.cell(row=i, column=2)
            w_cell.value = ""
        for i in range(2, sheet.max_row + 10):
            w_cell = sheet.cell(row=i, column=3)
            w_cell.value = ""
        for i in range(0, self.listWidget_subjects_db.count()):
            w_cell = sheet.cell(row=i + 2, column=3)
            w_cell.value = self.listWidget_subjects_db.item(i).text()
        for i in range(0, self.listWidget_lectors_db.count()):
            w_cell = sheet.cell(row=i + 2, column=1)
            w_cell.value = self.listWidget_lectors_db.item(i).text()
        for i in range(0, self.listWidget_assisstents_db.count()):
            w_cell = sheet.cell(row=i + 2, column=2)
            w_cell.value = self.listWidget_assisstents_db.item(i).text()
        workBookDB.save("resources\\DB_study_load.xlsx")
        workBookDB.close()
        self.load_db()

    def load_db(self):
        self.lectors = []
        self.assisstents = []
        self.subjects = []
        self.listWidget_subjects_db.clear()
        self.listWidget_assisstents_db.clear()
        self.listWidget_lectors_db.clear()
        self.comboBox_Subjects.clear()
        self.comboBox_Lectors.clear()
        self.comboBox_Assisstents.clear()
        workBookDB = load_workbook("resources\\DB_study_load.xlsx")
        sheet = workBookDB.active
        for i in range(2, sheet.max_row):
            if sheet.cell(column=1, row=i).value is not None:
                self.lectors.append(sheet.cell(column=1, row=i).value)
        for i in range(2, sheet.max_row):
            if sheet.cell(column=2, row=i).value is not None:
                self.assisstents.append(sheet.cell(column=2, row=i).value)
        for i in range(2, sheet.max_row):
            if sheet.cell(column=3, row=i).value is not None:
                self.subjects.append(sheet.cell(column=3, row=i).value)
        self.subjects.sort()
        self.assisstents.sort()
        self.lectors.sort()
        self.listWidget_assisstents_db.addItems(self.assisstents)
        self.listWidget_lectors_db.addItems(self.lectors)
        self.listWidget_subjects_db.addItems(self.subjects)
        self.comboBox_Assisstents.addItems(self.assisstents)
        self.comboBox_Lectors.addItems(self.lectors)
        self.comboBox_Subjects.addItems(self.subjects)
        workBookDB.close()

    def calculate_load(self):
        array_lectors = []
        for item in range(self.listWidget_lectors.count()):
            array_lectors.append(self.listWidget_lectors.item(item).text())
        array_assisstents = []
        for item in range(self.listWidget_assisstents.count()):
            array_assisstents.append(self.listWidget_assisstents.item(item).text())
        array_subject_count = []
        for item in range(self.listWidget_count_group.count()):
            array_subject_count.append(float(self.listWidget_count_group.item(item).text()))

        file_name = ""
        if self.comboBox_form_study.currentText() == "Сем І. Денна":
            file_name = self.default_file_names[self.comboBox.currentIndex()]
        elif self.comboBox_form_study.currentText() == "Сем ІІ. Денна":
            file_name = self.default_file_names[self.comboBox_2.currentIndex()]
        elif self.comboBox_form_study.currentText() == "Сем І. Заочна":
            file_name = self.default_file_names[self.comboBox_4.currentIndex()]
        elif self.comboBox_form_study.currentText() == "Сем ІІ. Заочна":
            file_name = self.default_file_names[self.comboBox_5.currentIndex()]

        try:
            nagr = Nagryzka02(file_name,
                            array_lectors,
                            array_assisstents,
                            len(array_lectors),
                            self.comboBox_Subjects.currentText(),
                            array_subject_count)
            info = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Information,
                                         "Розподіл навантаження", "Розподілення виконано успішно!",
                                         QtWidgets.QMessageBox.Ok)
            info.exec_()
        except:
            error = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning,
                                          "Розподіл навантаження", "Неправильна конфігурація!",
                                          QtWidgets.QMessageBox.Ok)
            error.exec_()

    def add_assisstant_db(self):
        if len(self.listWidget_assisstents_db.findItems(self.lineEdit_Assisstent.text(), Qt.Qt.MatchExactly)) == 0:
            self.listWidget_assisstents_db.addItem(self.lineEdit_Assisstent.text())
            self.lineEdit_Assisstent.clear()
            info = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Information,
                                         "Асистент", "Асистент був доданий до списку!",
                                         QtWidgets.QMessageBox.Ok)
            info.exec_()
        else:
            error = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning,
                                          "Асистент", "Такий вже є!",
                                          QtWidgets.QMessageBox.Ok)
            error.exec_()

    def add_lector_db(self):
        if len(self.listWidget_lectors_db.findItems(self.lineEdit_Lector.text(), Qt.Qt.MatchExactly)) == 0:
            self.listWidget_lectors_db.addItem(self.lineEdit_Lector.text())
            self.lineEdit_Lector.clear()
            info = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Information,
                                         "Лектор", "Лектор був доданий до списку!",
                                         QtWidgets.QMessageBox.Ok)
            info.exec_()
        else:
            error = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning,
                                          "Лектор", "Такий вже є!",
                                          QtWidgets.QMessageBox.Ok)
            error.exec_()

    def add_subject_db(self):
        if len(self.listWidget_subjects_db.findItems(self.lineEdit_predmet.text(), Qt.Qt.MatchExactly)) == 0:
            self.listWidget_subjects_db.addItem(self.lineEdit_predmet.text())
            self.lineEdit_predmet.clear()
            info = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Information,
                                         "Предмет", "Предмет був доданий до списку!",
                                         QtWidgets.QMessageBox.Ok)
            info.exec_()
        else:
            error = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning,
                                          "Предмет", "Такий вже є!",
                                          QtWidgets.QMessageBox.Ok)
            error.exec_()

    def remove_assisstant_db(self):
        self.listWidget_assisstents_db.takeItem(self.listWidget_assisstents_db.currentRow())

    def remove_lector_db(self):
        self.listWidget_lectors_db.takeItem(self.listWidget_lectors_db.currentRow())

    def remove_subject_db(self):
        self.listWidget_subjects_db.takeItem(self.listWidget_subjects_db.currentRow())

    def add_assisstant(self):
        self.listWidget_assisstents.addItem(self.comboBox_Assisstents.currentText())
        self.listWidget_count_group.addItem(self.comboBox_count_group.currentText())

    def add_lector(self):
        self.listWidget_lectors.addItem(self.comboBox_Lectors.currentText())

    def minus_assisstant(self):
        try:
            if self.listWidget_assisstents.currentItem().text() != '':
                self.listWidget_count_group.takeItem(self.listWidget_assisstents.currentRow())
                self.listWidget_assisstents.takeItem(self.listWidget_assisstents.currentRow())
        except AttributeError:
            error = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning,
                                          "Delete item", "You didn't select item", QtWidgets.QMessageBox.Ok)
            error.exec_()

    def minus_lector(self):
        try:
            self.listWidget_lectors.takeItem(self.listWidget_lectors.currentRow())
        except AttributeError:
            error = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning,
                                          "Delete item", "You didn't select item", QtWidgets.QMessageBox.Ok)
            error.exec_()

    def load_file(self):
        files_name = QtWidgets.QFileDialog.getOpenFileNames(self, "Open File", "resources", "XLS files (*.xls *.xlsx)")
        files_name = files_name[0]

        if len(files_name):
            try:
                self.comboBox.clear()
                self.comboBox_2.clear()
                self.comboBox_5.clear()
                self.comboBox_4.clear()
                self.comboBox.setEditable(False)
                self.comboBox_2.setEditable(False)
                self.comboBox_5.setEditable(False)
                self.comboBox_4.setEditable(False)
                self.only_file_name = []
                temporary_array = []
                for item in range(0, len(files_name)):
                    temporary_array.append(files_name[item].replace('/', '\\', 10))

                files_name = temporary_array
                self.default_file_names = files_name

                for item in range(0, len(files_name)):
                    temporary_array2 = files_name[item].split('\\')
                    self.only_file_name.append(temporary_array2[-1])
                self.comboBox.addItems(self.only_file_name)
                self.comboBox_2.addItems(self.only_file_name)
                self.comboBox_5.addItems(self.only_file_name)
                self.comboBox_4.addItems(self.only_file_name)
                self.comboBox.setCurrentIndex(0)
                self.comboBox_2.setCurrentIndex(2)
                self.comboBox_5.setCurrentIndex(3)
                self.comboBox_4.setCurrentIndex(1)

            except FileNotFoundError:
                error = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning,
                                              "Open Source File", "Failed to read file'", QtWidgets.QMessageBox.Ok)
                error.exec_()
            return

    def load_file_rnp(self):
        files_name = QtWidgets.QFileDialog.getOpenFileNames(self, "Open File", "resources", "XLS files (*.xls *.xlsx)")
        files_name = files_name[0]

        if len(files_name):
            try:
                self.comboBox_3.clear()
                self.comboBox_3.setEditable(False)
                self.only_file_name_rnp = []
                temporary_array = []
                for item in range(0, len(files_name)):
                    temporary_array.append(files_name[item].replace('/', '\\', 10))

                files_name = temporary_array
                self.default_file_names_rnp = files_name

                for item in range(0, len(files_name)):
                    temporary_array2 = files_name[item].split('\\')
                    self.only_file_name_rnp.append(temporary_array2[-1])
                self.comboBox_3.addItems(self.only_file_name_rnp)

            except FileNotFoundError:
                error = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning, "Open Source File",
                                              "Failed to read file\n", QtWidgets.QMessageBox.Ok)
                error.exec_()
            return

    def convert_file(self):
        try:
            ConverterXLS(self.default_file_names_rnp[self.comboBox_3.currentIndex()])
            self.default_file_names_rnp[self.comboBox_3.currentIndex()] += 'x'

            temp_array = list(self.comboBox_3.itemText(i) for i in range(self.comboBox_3.count()))
            print(self.default_file_names_rnp)

            temp_array[temp_array.index(self.comboBox_3.currentText())] += 'x'
            self.comboBox_3.clear()
            self.comboBox_3.addItems(temp_array)

        except IndexError:
            error = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning, "Convert File", "Failed to convert file\n",
                                          QtWidgets.QMessageBox.Ok)
            error.exec_()
        return

    def compare_files(self):
        try:
            if self.radioButton.isChecked():
                pars_object = pars(self.default_file_names_rnp[self.comboBox_3.currentIndex()],
                                   self.default_file_names[self.comboBox.currentIndex()],
                                   self.default_file_names[self.comboBox_2.currentIndex()])
            else:
                pars_object = pars(self.default_file_names_rnp[self.comboBox_3.currentIndex()],
                                   self.default_file_names[self.comboBox_4.currentIndex()],
                                   self.default_file_names[self.comboBox_5.currentIndex()])
            info = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Information, "Порінення файлів",
                                         "Порівнення файлів кафедри та РНП виконано успішно!!\n",
                                         QtWidgets.QMessageBox.Ok)
            info.exec_()
        except:
            error = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning, "Порівняння", "Неправильний формат файлу або файл вже відкритий!\n",
                                          QtWidgets.QMessageBox.Ok)
            error.exec_()
        return

    def create_report(self):
        try:
            save_name = QtWidgets.QFileDialog.getSaveFileName(self, "Save File", "resources",
                                                              "XLS files (*.xls *.xlsx)")
            report_Objcet = doc1(save_name, self.default_file_names[self.comboBox.currentIndex()],
                                 self.default_file_names[self.comboBox_2.currentIndex()],
                                 self.default_file_names[self.comboBox_4.currentIndex()],
                                 self.default_file_names[self.comboBox_5.currentIndex()])

            info = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Information, "Звіт",
                                         "Звіт був створений!\n",
                                         QtWidgets.QMessageBox.Ok)
            info.exec_()
        except:
            error = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning, "Звіт",
                                          "Відкрийте файли для створення звіту\n",
                                          QtWidgets.QMessageBox.Ok)
            error.exec_()
        return
