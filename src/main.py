from openpyxl import *

wb = load_workbook("resources\\ІТтаКБ. Сем I. Форма навчання  денна.xlsx")
sheet = wb.active
print(wb.get_sheet_names())
sheet.title
b = 0
c = 0
a = []

for i in range(18, 64):
    print(i, sheet.cell(row=i, column=2).value)

for j in range(6, 35):
    a.append(c)
    c = 0
    for i in range(18, 64):
     print(i, sheet.cell(row=i, column=j).value)
     if sheet.cell(row=i, column=j).value != 0 and sheet.cell(row=i, column=j).value != None and sheet.cell(row=i, column=j).value != "--" \
             and sheet.cell(row=i, column=j).value != "КП" and sheet.cell(row=i, column=j).value != "КР" and sheet.cell(row=i, column=j).value != "-"\
             and sheet.cell(row=i, column=j).value != "зал." and sheet.cell(row=i, column=j).value != "д.зал."  and sheet.cell(row=i, column=j).value != "усн.":
         if i != 54 and i != 55 and i != 56 and i != 57 and i != 58 and i != 60 and i != 61 and i != 62:
           c = c + (sheet.cell(row=i, column=j).value)




print(a)
ws = wb["Робота кафедри"]
wcell1 = ws.cell(20, 2)
wcell1.value = "PYTHON HERNYA"
wcell2 = ws.cell(21, 2)
wcell2.value = "Пайтон Херня"
wb.save("resources\\ІТтаКБ. Сем I. Форма навчання  денна.xlsx")
