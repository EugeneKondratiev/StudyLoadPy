from openpyxl import *


class ZaoсhnaSem2:
    def __init__(self):
        wb = load_workbook("resources\\ІТтаКБ. Сем II. Форма навчання  заочна.xlsx")
        sheet = wb.active
        print("Z2")
        print(wb.get_sheet_names())

        c = 0
        a = []
        i = 0
        j = 0

        for i in range(18, 64):
            print(i, sheet.cell(row=i, column=2).value)

        for j in range(6, 35):
            a.append(c)
            c = 0
            for i in range(18, 64):
                print(i, sheet.cell(row=i, column=j).value)
                if sheet.cell(row=i, column=j).value is not None \
                        and isinstance(sheet.cell(row=i, column=j).value, str) != True:
                    if i != 47 and i != 48 and i != 49 and i != 50 and i != 51 and i != 53 and i != 54 and i != 55 \
                            and i != 56 and i != 58 and i != 59 and i != 60 and i != 62 and i != 63 and i != 64:
                        c = c + sheet.cell(row=i, column=j).value

        print(a)
        wb.close()
