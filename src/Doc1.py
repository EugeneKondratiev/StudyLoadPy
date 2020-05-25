import self as self
from openpyxl import *

from src.DennaSem1 import DennaSem1
from src.DennaSem2 import DennaSem2
from src.ZaochnaSem2 import ZaoсhnaSem2
from src.ZaoсhnaSem1 import ZaoсhnaSem1


class doc1:
    #Считаем сумму общую за год, по семестрам, по стационару, по заочке, за семестр по видам деятельности (лекции, лабы и т.д.),
    # за год по видам деятельности. Далее это будет использоваться как контрольные суммы. Вид отчета: экран и файл *.ºxlsx (xls)
    denna1 = DennaSem1()
    den1 = denna1.a
    denna2 = DennaSem2()
    den2 = denna2.a
    zaochna1 = ZaoсhnaSem1()
    zaoch1 = zaochna1.a
    zaochna2 = ZaoсhnaSem2()
    zaoch2 = zaochna2.a

    zagal = [den1[28] + den2[28] + zaoch1[28] + zaoch2[28], den1[28] + den2[28], zaoch1[28] + zaoch2[28], den1[3] + den2[3] + zaoch1[3] + zaoch2[3],
             den1[11] + den2[11] + zaoch1[11] + zaoch2[11], den1[12] + den2[12] + zaoch1[12] + zaoch2[12], den1[13] + den2[13] + zaoch1[13] + zaoch2[13],
             den1[14] + den2[14] + zaoch1[14] + zaoch2[14], den1[15] + den2[15] + zaoch1[15] + zaoch2[15], den1[16] + den2[16] + zaoch1[16] + zaoch2[16],
             den1[17] + den2[17] + zaoch1[17] + zaoch2[17], den1[18] + den2[18] + zaoch1[18] + zaoch2[18], den1[19] + den2[19] + zaoch1[19] + zaoch2[19],
             den1[20] + den2[20] + zaoch1[20] + zaoch2[20], den1[21] + den2[21] + zaoch1[21] + zaoch2[21], den1[22] + den2[22] + zaoch1[22] + zaoch2[22],
             den1[23] + den2[23] + zaoch1[23] + zaoch2[23], den1[24] + den2[24] + zaoch1[24] + zaoch2[24], den1[25] + den2[25] + zaoch1[25] + zaoch2[25],
             den1[26] + den2[26] + zaoch1[26] + zaoch2[26], den1[27] + den2[27] + zaoch1[27] + zaoch2[27]]

    wb: Workbook = Workbook()
    ws = wb.active
    ws = wb["Sheet"]
    wcell1 = ws.cell(1, 1)
    wcell1.value = "Загальні данні"
    wcell2 = ws.cell(2, 1)
    wcell2.value = "За рік загальна"
    wcell3 = ws.cell(3, 1)
    wcell3.value = zagal[0]
    wcell4 = ws.cell(2, 2)
    wcell4.value = "Денна форма"
    wcell5 = ws.cell(3, 2)
    wcell5.value = zagal[1]
    wcell6 = ws.cell(2, 3)
    wcell6.value = "Заочна форма"
    wcell7 = ws.cell(3, 3)
    wcell7.value = zagal[2]
    wcell8 = ws.cell(2, 4)
    wcell8.value = "Всього кредитів ECTS"
    wcell9 = ws.cell(3, 4)
    wcell9.value = zagal[3]
    wcell10 = ws.cell(2, 5)
    wcell10.value = "Чит. лекцій"
    wcell11 = ws.cell(3, 5)
    wcell11.value = zagal[4]
    wcell12 = ws.cell(2, 6)
    wcell12.value = "Провед. лабор. занять"
    wcell13 = ws.cell(3, 6)
    wcell13.value = zagal[5]
    wcell14 = ws.cell(2, 7)
    wcell14.value = "Провед. практ./ семінар. занять"
    wcell15 = ws.cell(3, 7)
    wcell15.value = zagal[6]
    wcell16 = ws.cell(2, 8)
    wcell16.value = "Пров. консульт з нав. дисц. протягом семестру"
    wcell17 = ws.cell(3, 8)
    wcell17.value = zagal[7]
    wcell18 = ws.cell(2, 9)
    wcell18.value = "Пров. екзам. консультацій"
    wcell19 = ws.cell(3, 9)
    wcell19.value = zagal[8]
    wcell20 = ws.cell(2, 10)
    wcell20.value = "Керівництво і приймання КП/КР"
    wcell21 = ws.cell(3, 10)
    wcell21.value = zagal[9]
    wcell22 = ws.cell(2, 11)
    wcell22.value = "Пров. заліку"
    wcell23 = ws.cell(3, 11)
    wcell23.value = zagal[10]
    wcell24 = ws.cell(2, 12)
    wcell24.value = "Пров. сем. екз."
    wcell25 = ws.cell(3, 12)
    wcell25.value = zagal[11]
    wcell26 = ws.cell(2, 13)
    wcell26.value = "Кер-тво, консульт., реце-ня ДП"
    wcell27 = ws.cell(3, 13)
    wcell27.value = zagal[12]
    wcell28 = ws.cell(2, 14)
    wcell28.value = "Пров-ня захисту"
    wcell29 = ws.cell(3, 14)
    wcell29.value = zagal[13]
    wcell30 = ws.cell(2, 15)
    wcell30.value = "Кваліф. Іспит"
    wcell31 = ws.cell(3, 15)
    wcell31.value = zagal[14]
    wcell32 = ws.cell(2, 16)
    wcell32.value = "Кер-тво НДРС"
    wcell33 = ws.cell(3, 16)
    wcell33.value = zagal[15]
    wcell34 = ws.cell(2, 17)
    wcell34.value = "Кер-тво аспірантами, здобувачами"
    wcell35 = ws.cell(3, 17)
    wcell35.value = zagal[16]
    wcell36 = ws.cell(2, 18)
    wcell36.value = "Кер-тво практ"
    wcell37 = ws.cell(3, 18)
    wcell37.value = zagal[17]
    wcell38 = ws.cell(2, 19)
    wcell38.value = "Інші види -5%"
    wcell39 = ws.cell(3, 19)
    wcell39.value = zagal[18]
    wcell40 = ws.cell(2, 20)
    wcell40.value = "Дист. Модуль"
    wcell41 = ws.cell(3, 20)
    wcell41.value = zagal[19]
    wcell42 = ws.cell(2, 21)
    wcell42.value = "Додаткові години"
    wcell43 = ws.cell(3, 21)
    wcell43.value = zagal[20]
    wcell44 = ws.cell(2, 22)
    #По семестрам
    wb.create_sheet("По семестрам")
    ws = wb["По семестрам"]
    wcell1 = ws.cell(1, 1)
    wcell1.value = "Денна семестр 1"
    wcell2 = ws.cell(2, 1)
    wcell2.value = "К-сть груп"
    wcell3 = ws.cell(3, 1)
    wcell3.value = den1[1]
    wcell4 = ws.cell(2, 2)
    wcell4.value = "К-cть підгруп"
    wcell5 = ws.cell(3, 2)
    wcell5.value = den1[2]
    wcell6 = ws.cell(2, 3)
    wcell6.value = "К-cть потоків"
    wcell7 = ws.cell(3, 3)
    wcell7.value = den1[3]
    wcell8 = ws.cell(2, 4)
    wcell8.value = "Всього кредитів ECTS"
    wcell9 = ws.cell(3, 4)
    wcell9.value = den1[4]
    wcell10 = ws.cell(2, 5)
    wcell10.value = "Чит. лекцій"
    wcell11 = ws.cell(3, 5)
    wcell11.value = den1[11]
    wcell12 = ws.cell(2, 6)
    wcell12.value = "Провед. лабор. занять"
    wcell13 = ws.cell(3, 6)
    wcell13.value = den1[12]
    wcell14 = ws.cell(2, 7)
    wcell14.value = "Провед. практ./ семінар. занять"
    wcell15 = ws.cell(3, 7)
    wcell15.value = den1[13]
    wcell16 = ws.cell(2, 8)
    wcell16.value = "Пров. консульт з нав. дисц. протягом семестру"
    wcell17 = ws.cell(3, 8)
    wcell17.value = den1[14]
    wcell18 = ws.cell(2, 9)
    wcell18.value = "Пров. екзам. консультацій"
    wcell19 = ws.cell(3, 9)
    wcell19.value = den1[15]
    wcell20 = ws.cell(2, 10)
    wcell20.value = "Керівництво і приймання КП/КР"
    wcell21 = ws.cell(3, 10)
    wcell21.value = den1[16]
    wcell22 = ws.cell(2, 11)
    wcell22.value = "Пров. заліку"
    wcell23 = ws.cell(3, 11)
    wcell23.value = den1[17]
    wcell24 = ws.cell(2, 12)
    wcell24.value = "Пров. сем. екз."
    wcell25 = ws.cell(3, 12)
    wcell25.value = den1[18]
    wcell26 = ws.cell(2, 13)
    wcell26.value = "Кер-тво, консульт., реце-ня ДП"
    wcell27 = ws.cell(3, 13)
    wcell27.value = den1[19]
    wcell28 = ws.cell(2, 14)
    wcell28.value = "Пров-ня захисту"
    wcell29 = ws.cell(3, 14)
    wcell29.value = den1[20]
    wcell30 = ws.cell(2, 15)
    wcell30.value = "Кваліф. Іспит"
    wcell31 = ws.cell(3, 15)
    wcell31.value = den1[21]
    wcell32 = ws.cell(2, 16)
    wcell32.value = "Кер-тво НДРС"
    wcell33 = ws.cell(3, 16)
    wcell33.value = den1[22]
    wcell34 = ws.cell(2, 17)
    wcell34.value = "Кер-тво аспірантами, здобувачами"
    wcell35 = ws.cell(3, 17)
    wcell35.value = den1[23]
    wcell36 = ws.cell(2, 18)
    wcell36.value = "Кер-тво практ"
    wcell37 = ws.cell(3, 18)
    wcell37.value = den1[24]
    wcell38 = ws.cell(2, 19)
    wcell38.value = "Інші види -5%"
    wcell39 = ws.cell(3, 19)
    wcell39.value = den1[25]
    wcell40 = ws.cell(2, 20)
    wcell40.value = "Дист. Модуль"
    wcell41 = ws.cell(3, 20)
    wcell41.value = den1[26]
    wcell42 = ws.cell(2, 21)
    wcell42.value = "Додаткові години"
    wcell43 = ws.cell(3, 21)
    wcell43.value = den1[27]
    wcell44 = ws.cell(2, 22)
    wcell44.value = "Всього"
    wcell45 = ws.cell(3, 22)
    wcell45.value = den1[28]

    #denna sem2
    wcell1 = ws.cell(5, 1)
    wcell1.value = "Денна Семестр 2"
    wcell2 = ws.cell(6, 1)
    wcell2.value = "К-сть груп"
    wcell3 = ws.cell(7, 1)
    wcell3.value = den2[1]
    wcell4 = ws.cell(6, 2)
    wcell4.value = "К-cть підгруп"
    wcell5 = ws.cell(7, 2)
    wcell5.value = den2[2]
    wcell6 = ws.cell(6, 3)
    wcell6.value = "К-cть потоків"
    wcell7 = ws.cell(7, 3)
    wcell7.value = den2[3]
    wcell8 = ws.cell(6, 4)
    wcell8.value = "Всього кредитів ECTS"
    wcell9 = ws.cell(7, 4)
    wcell9.value = den2[4]
    wcell10 = ws.cell(6, 5)
    wcell10.value = "Чит. лекцій"
    wcell11 = ws.cell(7, 5)
    wcell11.value = den2[11]
    wcell12 = ws.cell(6, 6)
    wcell12.value = "Провед. лабор. занять"
    wcell13 = ws.cell(7, 6)
    wcell13.value = den2[12]
    wcell14 = ws.cell(6, 7)
    wcell14.value = "Провед. практ./ семінар. занять"
    wcell15 = ws.cell(7, 7)
    wcell15.value = den2[13]
    wcell16 = ws.cell(6, 8)
    wcell16.value = "Пров. консульт з нав. дисц. протягом семестру"
    wcell17 = ws.cell(7, 8)
    wcell17.value = den2[14]
    wcell18 = ws.cell(6, 9)
    wcell18.value = "Пров. екзам. консультацій"
    wcell19 = ws.cell(7, 9)
    wcell19.value = den2[15]
    wcell20 = ws.cell(6, 10)
    wcell20.value = "Керівництво і приймання КП/КР"
    wcell21 = ws.cell(7, 10)
    wcell21.value = den2[16]
    wcell22 = ws.cell(6, 11)
    wcell22.value = "Пров. заліку"
    wcell23 = ws.cell(7, 11)
    wcell23.value = den2[17]
    wcell24 = ws.cell(6, 12)
    wcell24.value = "Пров. сем. екз."
    wcell25 = ws.cell(7, 12)
    wcell25.value = den2[18]
    wcell26 = ws.cell(6, 13)
    wcell26.value = "Кер-тво, консульт., реце-ня ДП"
    wcell27 = ws.cell(7, 13)
    wcell27.value = den2[19]
    wcell28 = ws.cell(6, 14)
    wcell28.value = "Пров-ня захисту"
    wcell29 = ws.cell(7, 14)
    wcell29.value = den2[20]
    wcell30 = ws.cell(6, 15)
    wcell30.value = "Кваліф. Іспит"
    wcell31 = ws.cell(7, 15)
    wcell31.value = den2[21]
    wcell32 = ws.cell(6, 16)
    wcell32.value = "Кер-тво НДРС"
    wcell33 = ws.cell(7, 16)
    wcell33.value = den2[22]
    wcell34 = ws.cell(6, 17)
    wcell34.value = "Кер-тво аспірантами, здобувачами"
    wcell35 = ws.cell(7, 17)
    wcell35.value = den2[23]
    wcell36 = ws.cell(6, 18)
    wcell36.value = "Кер-тво практ"
    wcell37 = ws.cell(7, 18)
    wcell37.value = den2[24]
    wcell38 = ws.cell(6, 19)
    wcell38.value = "Інші види -5%"
    wcell39 = ws.cell(7, 19)
    wcell39.value = den2[25]
    wcell40 = ws.cell(6, 20)
    wcell40.value = "Дист. Модуль"
    wcell41 = ws.cell(7, 20)
    wcell41.value = den2[26]
    wcell42 = ws.cell(6, 21)
    wcell42.value = "Додаткові години"
    wcell43 = ws.cell(7, 21)
    wcell43.value = den2[27]
    wcell44 = ws.cell(6, 22)
    wcell44.value = "Всього"
    wcell45 = ws.cell(7, 22)
    wcell45.value = den2[28]
    # zaochna sem1
    wcell1 = ws.cell(9, 1)
    wcell1.value = "Заочна семестр 1"
    wcell2 = ws.cell(10, 1)
    wcell2.value = "К-сть груп"
    wcell3 = ws.cell(11, 1)
    wcell3.value = zaoch1[1]
    wcell4 = ws.cell(10, 2)
    wcell4.value = "К-cть підгруп"
    wcell5 = ws.cell(11, 2)
    wcell5.value = zaoch1[2]
    wcell6 = ws.cell(10, 3)
    wcell6.value = "К-cть потоків"
    wcell7 = ws.cell(11, 3)
    wcell7.value = zaoch1[3]
    wcell8 = ws.cell(10, 4)
    wcell8.value = "Всього кредитів ECTS"
    wcell9 = ws.cell(11, 4)
    wcell9.value = zaoch1[4]
    wcell10 = ws.cell(10, 5)
    wcell10.value = "Чит. лекцій"
    wcell11 = ws.cell(11, 5)
    wcell11.value = zaoch1[11]
    wcell12 = ws.cell(10, 6)
    wcell12.value = "Провед. лабор. занять"
    wcell13 = ws.cell(11, 6)
    wcell13.value = zaoch1[12]
    wcell14 = ws.cell(10, 7)
    wcell14.value = "Провед. практ./ семінар. занять"
    wcell15 = ws.cell(11, 7)
    wcell15.value = zaoch1[13]
    wcell16 = ws.cell(10, 8)
    wcell16.value = "Пров. консульт з нав. дисц. протягом семестру"
    wcell17 = ws.cell(11, 8)
    wcell17.value = zaoch1[14]
    wcell18 = ws.cell(10, 9)
    wcell18.value = "Пров. екзам. консультацій"
    wcell19 = ws.cell(11, 9)
    wcell19.value = zaoch1[15]
    wcell20 = ws.cell(10, 10)
    wcell20.value = "Керівництво і приймання КП/КР"
    wcell21 = ws.cell(11, 10)
    wcell21.value = zaoch1[16]
    wcell22 = ws.cell(10, 11)
    wcell22.value = "Пров. заліку"
    wcell23 = ws.cell(11, 11)
    wcell23.value = zaoch1[17]
    wcell24 = ws.cell(10, 12)
    wcell24.value = "Пров. сем. екз."
    wcell25 = ws.cell(11, 12)
    wcell25.value = zaoch1[18]
    wcell26 = ws.cell(10, 13)
    wcell26.value = "Кер-тво, консульт., реце-ня ДП"
    wcell27 = ws.cell(11, 13)
    wcell27.value = zaoch1[19]
    wcell28 = ws.cell(10, 14)
    wcell28.value = "Пров-ня захисту"
    wcell29 = ws.cell(11, 14)
    wcell29.value = zaoch1[20]
    wcell30 = ws.cell(10, 15)
    wcell30.value = "Кваліф. Іспит"
    wcell31 = ws.cell(11, 15)
    wcell31.value = zaoch1[21]
    wcell32 = ws.cell(10, 16)
    wcell32.value = "Кер-тво НДРС"
    wcell33 = ws.cell(11, 16)
    wcell33.value = zaoch1[22]
    wcell34 = ws.cell(10, 17)
    wcell34.value = "Кер-тво аспірантами, здобувачами"
    wcell35 = ws.cell(11, 17)
    wcell35.value = zaoch1[23]
    wcell36 = ws.cell(10, 18)
    wcell36.value = "Кер-тво практ"
    wcell37 = ws.cell(11, 18)
    wcell37.value = zaoch1[24]
    wcell38 = ws.cell(10, 19)
    wcell38.value = "Інші види -5%"
    wcell39 = ws.cell(11, 19)
    wcell39.value = zaoch1[25]
    wcell40 = ws.cell(10, 20)
    wcell40.value = "Дист. Модуль"
    wcell41 = ws.cell(11, 20)
    wcell41.value = zaoch1[26]
    wcell42 = ws.cell(10, 21)
    wcell42.value = "Додаткові години"
    wcell43 = ws.cell(11, 21)
    wcell43.value = zaoch1[27]
    wcell44 = ws.cell(10, 22)
    wcell44.value = "Всього"
    wcell45 = ws.cell(11, 22)
    wcell45.value = zaoch1[28]
    # zaochna sem2
    wcell1 = ws.cell(13, 1)
    wcell1.value = "Заочна семестр 2"
    wcell2 = ws.cell(14, 1)
    wcell2.value = "К-сть груп"
    wcell3 = ws.cell(15, 1)
    wcell3.value = zaoch2[1]
    wcell4 = ws.cell(14, 2)
    wcell4.value = "К-cть підгруп"
    wcell5 = ws.cell(15, 2)
    wcell5.value = zaoch2[2]
    wcell6 = ws.cell(14, 3)
    wcell6.value = "К-cть потоків"
    wcell7 = ws.cell(15, 3)
    wcell7.value = zaoch2[3]
    wcell8 = ws.cell(14, 4)
    wcell8.value = "Всього кредитів ECTS"
    wcell9 = ws.cell(15, 4)
    wcell9.value = zaoch2[4]
    wcell10 = ws.cell(14, 5)
    wcell10.value = "Чит. лекцій"
    wcell11 = ws.cell(15, 5)
    wcell11.value = zaoch2[11]
    wcell12 = ws.cell(14, 6)
    wcell12.value = "Провед. лабор. занять"
    wcell13 = ws.cell(15, 6)
    wcell13.value = zaoch2[12]
    wcell14 = ws.cell(14, 7)
    wcell14.value = "Провед. практ./ семінар. занять"
    wcell15 = ws.cell(15, 7)
    wcell15.value = zaoch2[13]
    wcell16 = ws.cell(14, 8)
    wcell16.value = "Пров. консульт з нав. дисц. протягом семестру"
    wcell17 = ws.cell(15, 8)
    wcell17.value = zaoch2[14]
    wcell18 = ws.cell(14, 9)
    wcell18.value = "Пров. екзам. консультацій"
    wcell19 = ws.cell(15, 9)
    wcell19.value = zaoch2[15]
    wcell20 = ws.cell(14, 10)
    wcell20.value = "Керівництво і приймання КП/КР"
    wcell21 = ws.cell(15, 10)
    wcell21.value = zaoch2[16]
    wcell22 = ws.cell(14, 11)
    wcell22.value = "Пров. заліку"
    wcell23 = ws.cell(15, 11)
    wcell23.value = zaoch2[17]
    wcell24 = ws.cell(14, 12)
    wcell24.value = "Пров. сем. екз."
    wcell25 = ws.cell(15, 12)
    wcell25.value = zaoch2[18]
    wcell26 = ws.cell(14, 13)
    wcell26.value = "Кер-тво, консульт., реце-ня ДП"
    wcell27 = ws.cell(15, 13)
    wcell27.value = zaoch2[19]
    wcell28 = ws.cell(14, 14)
    wcell28.value = "Пров-ня захисту"
    wcell29 = ws.cell(15, 14)
    wcell29.value = zaoch2[20]
    wcell30 = ws.cell(14, 15)
    wcell30.value = "Кваліф. Іспит"
    wcell31 = ws.cell(15, 15)
    wcell31.value = zaoch2[21]
    wcell32 = ws.cell(14, 16)
    wcell32.value = "Кер-тво НДРС"
    wcell33 = ws.cell(15, 16)
    wcell33.value = zaoch2[22]
    wcell34 = ws.cell(14, 17)
    wcell34.value = "Кер-тво аспірантами, здобувачами"
    wcell35 = ws.cell(15, 17)
    wcell35.value = zaoch2[23]
    wcell36 = ws.cell(14, 18)
    wcell36.value = "Кер-тво практ"
    wcell37 = ws.cell(15, 18)
    wcell37.value = zaoch2[24]
    wcell38 = ws.cell(14, 19)
    wcell38.value = "Інші види -5%"
    wcell39 = ws.cell(15, 19)
    wcell39.value = zaoch2[25]
    wcell40 = ws.cell(14, 20)
    wcell40.value = "Дист. Модуль"
    wcell41 = ws.cell(15, 20)
    wcell41.value = zaoch2[26]
    wcell42 = ws.cell(14, 21)
    wcell42.value = "Додаткові години"
    wcell43 = ws.cell(15, 21)
    wcell43.value = zaoch2[27]
    wcell44 = ws.cell(14, 22)
    wcell44.value = "Всього"
    wcell45 = ws.cell(15, 22)
    wcell45.value = zaoch2[28]
    wcell46 = ws.cell(14, 23)
    wb.save('Звіт1.xlsx')
    wb.close()