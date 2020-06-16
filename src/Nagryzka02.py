from openpyxl import *
from openpyxl.worksheet.worksheet import Worksheet
import numpy as np

class Nagryzka02:
    def __init__(self, file_name, PIB_lec, PIB_ass, count_lec, subject, subgroup_count_array):
        book = file_name
        wd = load_workbook(book)
        doc: Worksheet = wd.active
        piblek = PIB_lec
        pibass = PIB_ass
        podgrup_ass = subgroup_count_array
        countlek = count_lec
        predmet = subject
        chasy = []
        lektor = []
        ass = []
        kolvo_styd1 = 0
        grypp1 = ""
        potok1 = 0
        podgrup1 = 0
        den = {}
        c = []
        f_name = ""

        for i in range(18, doc.max_row):
            if doc.cell(row=i, column=2).value != None and doc.cell(row=i, column=2).value != 0:
                k = doc.cell(row=i - 1, column=2).value

                doc_cond = {temp: v for (temp, v) in den.items() if temp == k if temp != None}
                if len(doc_cond.items()) != 0 and t != "":
                    nympyarray1 = np.array(list(doc_cond.values()))
                    nympyarray2 = np.array(c)
                    nympyarray3 = np.array(nympyarray1 + nympyarray2)
                    den[t] = nympyarray3.tolist()
                else:
                    den[k] = c

                c = []
                t = k
                for j in range(16, doc.max_column + 1):
                    if doc.cell(row=i, column=j).value is not None and isinstance(doc.cell(row=i, column=j).value, str) != True:
                        c.append(doc.cell(row=i, column=j).value)

        del den[None]
        for i in range(1, doc.max_column):
            if predmet == doc.cell(row=i, column=2).value:
                grypp1 = doc.cell(row=i, column=5).value
                kolvo_styd1 = doc.cell(row=i, column=4).value
                podgrup1 = doc.cell(row=i, column=7).value
                potok1 = doc.cell(row=i, column=8).value

        for (k, c) in den.items():
            if predmet == k:
                chasy = c
        if "Сем I" in book and "денна" in book:
            f_name = "Sem1_Denna.xlsx"

        if "Сем I" in book and "заочна" in book:
            f_name = "Sem1_Zaochna.xlsx"

        if "Сем II" in book and "денна" in book:
            f_name = "Sem2_Denna.xlsx"

        if "Сем II" in book and "заочна" in book:
            f_name = "Sem2_Zaochna.xlsx"
        try:
            wt = load_workbook(f_name)
        except FileNotFoundError:
            wt: Workbook = Workbook()
        sheetnames = wt.sheetnames
        sheetequal = False
        for i in range(0, len(sheetnames)):
            if sheetnames[i] == predmet:
                sheetequal = True

        if sheetequal != True:
            wt.create_sheet(predmet[:30])
        wc = wt[predmet]
        for i in range(0, len(chasy)):
            if i == 0 or i == 3 or i == 4 or i == 5 or i == 6 or i == 7 or i == 8 or i == 9 or i == 10 or i == 11 or i == 12 or i == 13 or i == 15:
                lektor.append(float(chasy[i] / countlek))
                ass.append(0)
            elif i != 17:
                ass.append(float(chasy[i] / podgrup1))
                lektor.append(0)

        if len(ass) != 0:
            wcell1 = wc.cell(3, 1)
            wcell1.value = "ПІБ викладача"
            wcell4 = wc.cell(3, 2)
            wcell4.value = "К-сть студ"
            wcell4 = wc.cell(3, 3)
            wcell4.value = "Шифр групп"
            wcell4 = wc.cell(3, 4)
            wcell4.value = "К-сть Потоків"
            wcell4 = wc.cell(3, 5)
            wcell4.value = "К-сть Підгруп"
            wcell2 = wc.cell(3, 6)
            wcell2.value = "Чит. лекцій"
            wcell1 = wc.cell(3, 7)
            wcell1.value = "Провед. лабор. занять"
            wcell1 = wc.cell(3, 8)
            wcell1.value = "Провед. практ./ семінар. занять"
            wcell4 = wc.cell(3, 9)
            wcell4.value = "Пров. консульт з нав. дисц. протягом семестру"
            wcell6 = wc.cell(3, 10)
            wcell6.value = "Пров. екзам. консультацій"
            wcell8 = wc.cell(3, 11)
            wcell8.value = "Керівництво і приймання КП/КР"
            wcell10 = wc.cell(3, 12)
            wcell10.value = "Пров. заліку"
            wcell12 = wc.cell(3, 13)
            wcell12.value = "Пров. сем. екз."
            wcell14 = wc.cell(3, 14)
            wcell14.value = "Кер-тво, консульт., реце-ня ДП"
            wcell4 = wc.cell(3, 15)
            wcell4.value = "Пров-ня захисту"
            wcell4 = wc.cell(3, 16)
            wcell4.value = "Кваліф. Іспит"
            wcell4 = wc.cell(3, 17)
            wcell4.value = "Кер-тво НДРС"
            wcell4 = wc.cell(3, 18)
            wcell4.value = "Кер-тво аспірантами, здобувачами"
            wcell4 = wc.cell(3, 19)
            wcell4.value = "Кер-тво практ."
            wcell1 = wc.cell(3, 20)
            wcell1.value = "Інші види -5%"
            wcell4 = wc.cell(3, 21)
            wcell4.value = "Дист. Модуль"
            wcell1 = wc.cell(3, 22)
            wcell1.value = "Додаткові години"
            wcell1 = wc.cell(3, 23)
            wcell1.value = "Всього"

            #Название предмета
            wcell1 = wc.cell(1, 1)
            wcell1.value = predmet

            wcell1 = wc.cell(2, 1)
            wcell1.value = "Семестр1"
            rows = 4

            #ФИО лекторов
            for i in range(0, len(piblek)):
                wcell2 = wc.cell(rows, 1)
                wcell2.value = piblek[i]
                rows = rows + 1

            #ФИО асистентов
            for i in range(0, len(pibass)):
                wcell3 = wc.cell(rows, 1)
                wcell3.value = pibass[i]
                rows = rows + 1


            rows = 4
            colums = 2
            # Часы лекторов
            for i in range(0, len(piblek)):
                q = 0
                colums = 2
                wcell6 = wc.cell(rows, colums)
                wcell6.value = kolvo_styd1
                colums = colums + 1
                wcell7 = wc.cell(rows, colums)
                wcell7.value = grypp1
                colums = colums + 1
                wcell7 = wc.cell(rows, colums)
                wcell7.value = potok1
                colums = colums + 1
                wcell6 = wc.cell(rows, colums)
                wcell6.value = podgrup1
                colums = colums + 1
                for j in range(0, len(lektor)):
                    wcell4 = wc.cell(rows, colums)
                    wcell4.value = lektor[j]
                    q = q + lektor[j]
                    colums = colums + 1
                wcell6 = wc.cell(rows, colums)
                wcell6.value = q
                rows = rows + 1

            colums = 2
            # Часы асистентов
            for i in range(0, len(pibass)):
                q = 0
                colums = 2
                wcell6 = wc.cell(rows, colums)
                wcell6.value = kolvo_styd1
                colums = colums + 1
                wcell7 = wc.cell(rows, colums)
                wcell7.value = grypp1
                colums = colums + 1
                wcell7 = wc.cell(rows, colums)
                wcell7.value = potok1
                colums = colums + 1
                wcell6 = wc.cell(rows, colums)
                wcell6.value = podgrup_ass[i]
                colums = colums + 1
                for j in range(0, len(ass)):
                    wcell5 = wc.cell(rows, colums)
                    wcell5.value = ass[j] * podgrup_ass[i]
                    q = q + ass[j] * podgrup_ass[i]
                    colums = colums + 1
                wcell6 = wc.cell(rows, colums)
                wcell6.value = q
                rows = rows + 1

            colums = 1
            wcell6 = wc.cell(rows, colums)
            wcell6.value = "Всього"
            colums = colums + 1
            wcell6 = wc.cell(rows, colums)
            wcell6.value = kolvo_styd1
            colums = colums + 1
            wcell7 = wc.cell(rows, colums)
            wcell7.value = grypp1
            colums = colums + 1
            wcell7 = wc.cell(rows, colums)
            wcell7.value = potok1
            colums = colums + 1
            wcell6 = wc.cell(rows, colums)
            wcell6.value = podgrup1
            colums = colums + 1
            for j in range(0, len(chasy)):
                wcell5 = wc.cell(rows, colums)
                wcell5.value = chasy[j]
                colums = colums + 1

        for item in wt.sheetnames:
            if item == 'Sheet':
                empty_sheet = wt.get_sheet_by_name('Sheet')
                wt.remove_sheet(empty_sheet)

        wt.save(f_name)
        wt.close()
        wd.close()
