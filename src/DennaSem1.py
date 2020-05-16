from openpyxl import *


class DennaSem1:
    def __init__(self):
        wb = load_workbook("resources\\ІТтаКБ. Сем I. Форма навчання  денна.xlsx")
        sheet = wb.active
        print("DENNA")
        print(wb.get_sheet_names())

        c = 0
        a = []
        i = 0
        j = 0

        for i in range(18, 64):
            print(i, sheet.cell(row=i, column=2).value)

        for j in range(6, 35):
            a.append(c)
            c = 0
            for i in range(18, 64):
                print(i, sheet.cell(row=i, column=j).value)
                if sheet.cell(row=i, column=j).value is not None \
                        and isinstance(sheet.cell(row=i, column=j).value, str) != True:
                    if i != 54 and i != 55 and i != 56 and i != 57 and i != 58 and i != 60 and i != 61 and i != 62:
                        c = c + sheet.cell(row=i, column=j).value

        print(a)
        wb.close()
