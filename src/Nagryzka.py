from openpyxl import *
from openpyxl.worksheet.worksheet import Worksheet
import numpy as np


class Nagryzka:
    def __init__(self, file_name_1sem_den, file_name_1sem_zaoch, file_name_2sem_den, file_name_2sem_zaoch, PIB_lec,
                 PIB_ass, count_lec, count_ass, subject, subgroup_count_array):
        wd = load_workbook(file_name_1sem_den)
        ws = load_workbook(file_name_1sem_zaoch)
        wq = load_workbook(file_name_2sem_den)
        ww = load_workbook(file_name_2sem_zaoch)
        denna2: Worksheet = wq.active
        zaochna2: Worksheet = ww.active
        denna1: Worksheet = wd.active
        zaochna1: Worksheet = ws.active
        piblek = PIB_lec
        pibass = PIB_ass
        podgrup1 = 0
        podgrup2 = 0
        podgrup_ass = subgroup_count_array
        countlek = count_lec
        countass = count_ass
        predmet = subject
        chasy = []
        chasy2 = []
        lektor = []
        lektor2 = []
        ass = []
        ass2 = []
        kolvo_styd1 = 0
        kolvo_styd2 = 0
        grypp1 = ""
        grypp2 = ""

        den = {}
        zaoch = {}
        c = []
        for i in range(18, denna1.max_row):
            if denna1.cell(row=i, column=2).value != None and denna1.cell(row=i, column=2).value != 0:
                k = denna1.cell(row=i - 1, column=2).value

                dict1_cond = {temp: v for (temp, v) in den.items() if temp == k if temp != None}
                if len(dict1_cond.items()) != 0 and t != "":
                    nympyarray1 = np.array(list(dict1_cond.values()))
                    nympyarray2 = np.array(c)
                    nympyarray3 = np.array(nympyarray1 + nympyarray2)
                    den[t] = nympyarray3.tolist()
                else:
                    den[k] = c

                c = []
                t = k
                for j in range(16, denna1.max_column + 1):
                    if denna1.cell(row=i, column=j).value is not None and isinstance(denna1.cell(row=i, column=j).value,
                                                                                     str) != True:
                        c.append(denna1.cell(row=i, column=j).value)
        c = []
        t = ""
        y = 0
        for i in range(18, zaochna1.max_row):
            if zaochna1.cell(row=i, column=2).value != None and zaochna1.cell(row=i, column=2).value != "":
                k = zaochna1.cell(row=i - 1, column=2).value
                dict1_cond = {temp: v for (temp, v) in zaoch.items() if temp == k if temp != None}
                if len(dict1_cond.items()) != 0 and t != "":
                    nympyarray1 = np.array(list(dict1_cond.values()))
                    nympyarray2 = np.array(c)
                    nympyarray3 = np.array(nympyarray1 + nympyarray2)
                    zaoch[t] = nympyarray3.tolist()
                else:
                    zaoch[k] = c
                t = k
                c = []
                for j in range(16, zaochna1.max_column + 1):
                    if zaochna1.cell(row=i, column=j).value is not None and isinstance(
                            zaochna1.cell(row=i, column=j).value, str) != True:
                        c.append(zaochna1.cell(row=i, column=j).value)
        del den[None]
        del zaoch[None]
        dict1_tripleCond = {}
        for (temp, v) in den.items():
            ikval = False
            for (k, c) in zaoch.items():
                if temp == k and temp != None:
                    for i in range(1, denna1.max_column):
                        if predmet == denna1.cell(row=i, column=2).value and k == predmet:
                            grypp1 = denna1.cell(row=i, column=5).value + zaochna1.cell(row=i, column=5).value
                        if predmet == denna1.cell(row=i, column=2).value and k == predmet:
                            kolvo_styd1 = denna1.cell(row=i, column=4).value + zaochna1.cell(row=i, column=4).value
                        if predmet == denna1.cell(row=i, column=2).value and k == predmet:
                            podgrup1 = denna1.cell(row=i, column=7).value + zaochna1.cell(row=i, column=7).value
                    nympyarray1 = np.array(v)
                    nympyarray2 = np.array(c)
                    nympyarray3 = np.array(nympyarray1 + nympyarray2)
                    dict1_tripleCond[k] = nympyarray3.tolist()
                    ikval = True
            if ikval== False:
                dict1_tripleCond[temp] = v
                dict1_tripleCond[k] = c
        for i in range(1, denna1.max_column):
            if predmet == denna1.cell(row=i, column=2).value:
                grypp1 = denna1.cell(row=i, column=5).value
            elif predmet == zaochna1.cell(row=i, column=2).value:
                grypp1 = zaochna1.cell(row=i, column=5).value

            if predmet == denna1.cell(row=i, column=2).value:
                kolvo_styd1 = denna1.cell(row=i, column=4).value
            elif predmet == zaochna1.cell(row=i, column=2).value:
                kolvo_styd1 = zaochna1.cell(row=i, column=4).value

            if predmet == denna1.cell(row=i, column=2).value:
                podgrup1 = denna1.cell(row=i, column=7).value
            elif predmet == zaochna1.cell(row=i, column=2).value:
                podgrup1 = zaochna1.cell(row=i, column=7).value

        den2 = {}
        zaoch2 = {}
        c = []
        k = ""
        for i in range(18, denna2.max_row):
            if denna2.cell(row=i, column=2).value != None and denna2.cell(row=i, column=2).value != 0:
                k = denna2.cell(row=i - 1, column=2).value

                dict2_cond = {temp: v for (temp, v) in den.items() if temp == k if temp != None}
                if len(dict2_cond.items()) != 0 and t != "":
                    nympyarray1 = np.array(list(dict2_cond.values()))
                    nympyarray2 = np.array(c)
                    nympyarray3 = np.array(nympyarray1 + nympyarray2)
                    den2[t] = nympyarray3.tolist()
                else:
                    den2[k] = c

                c = []
                t = k
                for j in range(16, denna2.max_column + 1):
                    if denna2.cell(row=i, column=j).value is not None and isinstance(denna2.cell(row=i, column=j).value,
                                                                                     str) != True:
                        c.append(denna2.cell(row=i, column=j).value)
        c = []
        t = ""
        y = 0
        for i in range(18, zaochna2.max_row):
            if zaochna2.cell(row=i, column=2).value != None and zaochna2.cell(row=i, column=2).value != "":
                k = zaochna2.cell(row=i - 1, column=2).value
                dict2_cond = {temp: v for (temp, v) in zaoch.items() if temp == k if temp != None}
                if len(dict2_cond.items()) != 0 and t != "":
                    nympyarray1 = np.array(list(dict2_cond.values()))
                    nympyarray2 = np.array(c)
                    nympyarray3 = np.array(nympyarray1 + nympyarray2)
                    zaoch2[t] = nympyarray3.tolist()
                else:
                    zaoch2[k] = c
                t = k
                c = []
                for j in range(16, zaochna2.max_column + 1):
                    if zaochna2.cell(row=i, column=j).value is not None and isinstance(
                            zaochna2.cell(row=i, column=j).value, str) != True:
                        c.append(zaochna2.cell(row=i, column=j).value)

        del den2[None]
        del zaoch2[None]
        dict2_tripleCond = {}
        for (temp, v) in den2.items():
            ikval = False
            for (k, c) in zaoch2.items():
                if temp == k and temp != None:
                    for i in range(1, denna2.max_column):
                        if predmet == denna2.cell(row=i, column=2).value and k == predmet:
                            grypp2 = denna2.cell(row=i, column=7).value + zaochna2.cell(row=i, column=7).value
                        if predmet == denna2.cell(row=i, column=2).value and k == predmet:
                            kolvo_styd2 = denna2.cell(row=i, column=7).value + zaochna2.cell(row=i, column=7).value
                        if predmet == denna2.cell(row=i, column=2).value and k == predmet:
                            podgrup2 = denna2.cell(row=i, column=7).value + zaochna2.cell(row=i, column=7).value
                    nympyarray1 = np.array(v)
                    nympyarray2 = np.array(c)
                    nympyarray3 = np.array(nympyarray1 + nympyarray2)
                    dict2_tripleCond[k] = nympyarray3.tolist()
                    ikval = True
            if ikval== False:
                dict2_tripleCond[temp] = v
                dict2_tripleCond[k] = c

        for i in range(1, denna2.max_column):
            if predmet == denna2.cell(row=i, column=2).value:
                grypp2 = denna2.cell(row=i, column=5).value
            elif predmet == zaochna2.cell(row=i, column=2).value:
                grypp2 = zaochna2.cell(row=i, column=5).value

            if predmet == denna2.cell(row=i, column=2).value:
                kolvo_styd2 = denna2.cell(row=i, column=4).value
            elif predmet == zaochna2.cell(row=i, column=2).value:
                kolvo_styd2 = zaochna2.cell(row=i, column=4).value

            if predmet == denna2.cell(row=i, column=2).value:
                podgrup2 = denna2.cell(row=i, column=7).value
            elif predmet == zaochna2.cell(row=i, column=2).value:
                podgrup2 = zaochna2.cell(row=i, column=7).value

        #НАЧАЛО КОНЦА


        for (k, c) in dict1_tripleCond.items():
            if predmet == k:
                chasy = c

        k = ""
        for (y, c) in dict2_tripleCond.items():
            if predmet == y:
                chasy2 = c

        try:
            wt = load_workbook("Розподіл навантаження.xlsx")
        except FileNotFoundError:
            wt: Workbook = Workbook()
        sheetnames = wt.sheetnames
        sheetequal = False
        for i in range (0, len(sheetnames)):
            if sheetnames[i] == predmet:
                sheetequal = True

        if sheetequal != True:
            wt.create_sheet(predmet[:30])
        wc = wt[predmet]


        for i in range(0, len(chasy)):
            if i == 0 or i == 3 or i == 4 or i == 5 or i == 6 or i == 7 or i == 8 or i == 9 or i == 10 or i == 11 or i == 12 or i == 13 or i == 15:
                lektor.append(chasy[i] / countlek)
            elif i != 17:
                ass.append(chasy[i] / podgrup1)

        for i in range(0, len(chasy2)):
            if i == 0 or i == 3 or i == 4 or i == 5 or i == 6 or i == 7 or i == 8 or i == 9 or i == 10 or i == 11 or i == 12 or i == 13 or i == 15:
                lektor2.append(chasy2[i] / countlek)
            elif i != 17:
                ass2.append(chasy2[i] / podgrup2)
        if len(ass2) != 0:
            wcell1 = wc.cell(13, 1)
            wcell1.value = "Семестр 2 "
            wcell1 = wc.cell(14, 1)
            wcell1.value = "ПІБ викладача"
            wcell2 = wc.cell(14, 2)
            wcell2.value = "Чит. лекцій"
            wcell4 = wc.cell(14, 3)
            wcell4.value = "Пров. консульт з нав. дисц. протягом семестру"
            wcell6 = wc.cell(14, 4)
            wcell6.value = "Пров. екзам. консультацій"
            wcell8 = wc.cell(14, 5)
            wcell8.value = "Керівництво і приймання КП/КР"
            wcell10 = wc.cell(14, 6)
            wcell10.value = "Пров. заліку"
            wcell12 = wc.cell(14, 7)
            wcell12.value = "Пров. сем. екз."
            wcell14 = wc.cell(14, 8)
            wcell14.value = "Кер-тво, консульт., реце-ня ДП"
            wcell4 = wc.cell(14, 9)
            wcell4.value = "Пров-ня захисту"
            wcell4 = wc.cell(14, 10)
            wcell4.value = "Кваліф. Іспит"
            wcell4 = wc.cell(14, 11)
            wcell4.value = "Кер-тво НДРС"
            wcell4 = wc.cell(14, 12)
            wcell4.value = "Кер-тво аспірантами, здобувачами"
            wcell4 = wc.cell(14, 13)
            wcell4.value = "Кер-тво практ"
            wcell4 = wc.cell(14, 14)
            wcell4.value = "Дист. Модуль"
            wcell4 = wc.cell(14, 15)
            wcell4.value = "К-сть студ"
            wcell4 = wc.cell(14, 16)
            wcell4.value = "Шифр групп"
            wcell1 = wc.cell(16, 1)
            wcell1.value = "ПІБ викладача"
            wcell1 = wc.cell(16, 2)
            wcell1.value = "Провед. лабор. занять"
            wcell1 = wc.cell(16, 3)
            wcell1.value = "Провед. практ./ семінар. занять"
            wcell1 = wc.cell(16, 4)
            wcell1.value = "Інші види -5%"
            wcell1 = wc.cell(16, 5)
            wcell1.value = "Додаткові години"
            wcell1 = wc.cell(16, 6)
            wcell1.value = "Кількісль Підгруп."

            rows = 15
            # ФИО лекторов
            for i in range(0, len(piblek)):
             wcell2 = wc.cell(rows, 1)
             wcell2.value = piblek[i]
             rows = rows + 1

            # ФИО асистентов
            rows = rows + 1
            for i in range(0, len(pibass)):
             wcell3 = wc.cell(rows, 1)
             wcell3.value = pibass[i]
             rows = rows + 1

            rows = 15
            colums = 2
            # Часы лекторов
            for i in range(0, len(piblek)):
                colums = 2
                for j in range(0, len(lektor2)):
                    wcell4 = wc.cell(rows, colums)
                    wcell4.value = lektor2[j]
                    colums = colums + 1
                wcell6 = wc.cell(rows, colums)
                wcell6.value = kolvo_styd2
                colums = colums + 1
                wcell7 = wc.cell(rows, colums)
                wcell7.value = grypp2
                rows = rows + 1

            colums = 2
            rows = rows + 1
            # Часы асистентов
            for i in range(0, len(pibass)):
                colums = 2
                for j in range(0, len(ass2)):
                    wcell5 = wc.cell(rows, colums)
                    wcell5.value = ass2[j] * podgrup_ass[i]
                    colums = colums + 1
                wcell6 = wc.cell(rows, colums)
                wcell6.value = podgrup_ass[i]
                rows = rows + 1

        if len(ass) != 0:
            wcell1 = wc.cell(3, 1)
            wcell1.value = "ПІБ викладача"
            wcell2 = wc.cell(3, 2)
            wcell2.value = "Чит. лекцій"
            wcell4 = wc.cell(3, 3)
            wcell4.value = "Пров. консульт з нав. дисц. протягом семестру"
            wcell6 = wc.cell(3, 4)
            wcell6.value = "Пров. екзам. консультацій"
            wcell8 = wc.cell(3, 5)
            wcell8.value = "Керівництво і приймання КП/КР"
            wcell10 = wc.cell(3, 6)
            wcell10.value = "Пров. заліку"
            wcell12 = wc.cell(3, 7)
            wcell12.value = "Пров. сем. екз."
            wcell14 = wc.cell(3, 8)
            wcell14.value = "Кер-тво, консульт., реце-ня ДП"
            wcell4 = wc.cell(3, 9)
            wcell4.value = "Пров-ня захисту"
            wcell4 = wc.cell(3, 10)
            wcell4.value = "Кваліф. Іспит"
            wcell4 = wc.cell(3, 11)
            wcell4.value = "Кер-тво НДРС"
            wcell4 = wc.cell(3, 12)
            wcell4.value = "Кер-тво аспірантами, здобувачами"
            wcell4 = wc.cell(3, 13)
            wcell4.value = "Кер-тво практ."
            wcell4 = wc.cell(3, 14)
            wcell4.value = "Дист. Модуль"
            wcell4 = wc.cell(3, 15)
            wcell4.value = "К-сть студ"
            wcell4 = wc.cell(3, 16)
            wcell4.value = "Шифр групп"
            wcell1 = wc.cell(5, 1)
            wcell1.value = "ПІБ викладача"
            wcell1 = wc.cell(5, 2)
            wcell1.value = "Провед. лабор. занять"
            wcell1 = wc.cell(5, 3)
            wcell1.value = "Провед. практ./ семінар. занять"
            wcell1 = wc.cell(5, 4)
            wcell1.value = "Інші види -5%"
            wcell1 = wc.cell(5, 5)
            wcell1.value = "Додаткові години"
            wcell1 = wc.cell(5, 6)
            wcell1.value = "Кількісль Підгруп."
            #Название предмета
            wcell1 = wc.cell(1, 1)
            wcell1.value = predmet

            wcell1 = wc.cell(2, 1)
            wcell1.value = "Семестр1"
            rows = 4

            #ФИО лекторов
            for i in range(0, len(piblek)):
                wcell2 = wc.cell(rows, 1)
                wcell2.value = piblek[i]
                rows = rows + 1

            #ФИО асистентов
            rows = rows + 1
            for i in range(0, len(pibass)):
                wcell3 = wc.cell(rows, 1)
                wcell3.value = pibass[i]
                rows = rows + 1


            rows = 4
            colums = 2
            # Часы лекторов
            for i in range(0, len(piblek)):
                colums = 2
                for j in range(0, len(lektor)):
                    wcell4 = wc.cell(rows, colums)
                    wcell4.value = lektor[j]
                    colums = colums + 1
                wcell6 = wc.cell(rows, colums)
                wcell6.value = kolvo_styd1
                colums = colums + 1
                wcell7 = wc.cell(rows, colums)
                wcell7.value = grypp1
                rows = rows + 1

            colums = 2
            rows = rows + 1
            # Часы асистентов
            for i in range(0, len(pibass)):
                colums = 2
                for j in range(0, len(ass)):
                    wcell5 = wc.cell(rows, colums)
                    wcell5.value = int(ass[j]) * int(podgrup_ass[i])
                    colums = colums + 1
                wcell6 = wc.cell(rows, colums)
                wcell6.value = podgrup_ass[i]
                rows = rows + 1


        wt.save("Розподіл навантаження.xlsx")
        wt.close()
        wd.close()
        ws.close()
