from openpyxl import *
from openpyxl.worksheet.worksheet import Worksheet
import numpy as np


class Zvit3:
    def __init__(self):
        wb = load_workbook("Sem1_Denna.xlsx")
        doc1: Worksheet = wb.active
        wq = load_workbook("Sem2_Denna.xlsx")
        doc2: Worksheet = wq.active
        wz = load_workbook("Sem1_Zaochna.xlsx")
        doc3: Worksheet = wz.active
        wr = load_workbook("Sem2_Zaochna.xlsx")
        doc4: Worksheet = wr.active
        wd = load_workbook("resources\\DB_study_load.xlsx")
        fam: Worksheet = wd.active
        try:
            wt = load_workbook("zvit2.xlsx")
        except FileNotFoundError:
            wt: Workbook = Workbook()
        k = ""




        for i in range(1, fam.max_row):
            if fam.cell(row=i, column=2).value != None:
                k = fam.cell(row=i, column=2).value
                p = 0
            rows = 2
            for sheet in wb:
                wb.active = sheet
                doc1: Worksheet = wb.active

                for j in range(1, doc1.max_row + 1):
                    if k == doc1.cell(row=j, column=1).value:
                        sheetnames = wt.sheetnames
                        sheetequal = False
                        for l in range(0, len(sheetnames)):
                            if sheetnames[l] == k:
                                sheetequal = True
                        if sheetequal != True:
                            wt.create_sheet(k[:30])
                        wc = wt[k]
                        p = p + 1
                        if p == 1:
                            wcell1 = wc.cell(1, 1)
                            wcell1.value = k
                            wcell1 = wc.cell(rows-1, 5)
                            wcell1.value = "Семестр 1 Денна"
                            wcell1 = wc.cell(rows, 1)
                            wcell1.value = "Назва навчальних дисциплін"
                            wcell4 = wc.cell(rows, 2)
                            wcell4.value = "К-сть студ"
                            wcell4 = wc.cell(rows, 3)
                            wcell4.value = "Шифр групп"
                            wcell4 = wc.cell(rows, 4)
                            wcell4.value = "К-сть Потоків"
                            wcell4 = wc.cell(rows, 5)
                            wcell4.value = "К-сть Підгруп"
                            wcell2 = wc.cell(rows, 6)
                            wcell2.value = "Чит. лекцій"
                            wcell1 = wc.cell(rows, 7)
                            wcell1.value = "Провед. лабор. занять"
                            wcell1 = wc.cell(rows, 8)
                            wcell1.value = "Провед. практ./ семінар. занять"
                            wcell4 = wc.cell(rows, 9)
                            wcell4.value = "Пров. консульт з нав. дисц. протягом семестру"
                            wcell6 = wc.cell(rows, 10)
                            wcell6.value = "Пров. екзам. консультацій"
                            wcell8 = wc.cell(rows, 11)
                            wcell8.value = "Керівництво і приймання КП/КР"
                            wcell10 = wc.cell(rows, 12)
                            wcell10.value = "Пров. заліку"
                            wcell12 = wc.cell(rows, 13)
                            wcell12.value = "Пров. сем. екз."
                            wcell14 = wc.cell(rows, 14)
                            wcell14.value = "Кер-тво, консульт., реце-ня ДП"
                            wcell4 = wc.cell(rows, 15)
                            wcell4.value = "Пров-ня захисту"
                            wcell4 = wc.cell(rows, 16)
                            wcell4.value = "Кваліф. Іспит"
                            wcell4 = wc.cell(rows, 17)
                            wcell4.value = "Кер-тво НДРС"
                            wcell4 = wc.cell(rows, 18)
                            wcell4.value = "Кер-тво аспірантами, здобувачами"
                            wcell4 = wc.cell(rows, 19)
                            wcell4.value = "Кер-тво практ."
                            wcell1 = wc.cell(rows, 20)
                            wcell1.value = "Інші види -5%"
                            wcell4 = wc.cell(rows, 21)
                            wcell4.value = "Дист. Модуль"
                            wcell1 = wc.cell(rows, 22)
                            wcell1.value = "Додаткові години"
                            wcell1 = wc.cell(rows, 23)
                            wcell1.value = "Всього"
                            rows = rows + 1

                        wcell1 = wc.cell(rows, 1)
                        wcell1.value = doc1.cell(row=1, column=1).value
                        wcell1 = wc.cell(rows, 2)
                        wcell1.value = doc1.cell(row=j, column=2).value
                        wcell1 = wc.cell(rows, 3)
                        wcell1.value = doc1.cell(row=j, column=3).value
                        wcell1 = wc.cell(rows, 4)
                        wcell1.value = doc1.cell(row=j, column=4).value
                        wcell1 = wc.cell(rows, 5)
                        wcell1.value = doc1.cell(row=j, column=5).value
                        wcell1 = wc.cell(rows, 6)
                        wcell1.value = doc1.cell(row=j, column=6).value
                        wcell1 = wc.cell(rows, 7)
                        wcell1.value = doc1.cell(row=j, column=7).value
                        wcell1 = wc.cell(rows, 8)
                        wcell1.value = doc1.cell(row=j, column=8).value
                        wcell1 = wc.cell(rows, 9)
                        wcell1.value = doc1.cell(row=j, column=9).value
                        wcell1 = wc.cell(rows, 10)
                        wcell1.value = doc1.cell(row=j, column=10).value
                        wcell1 = wc.cell(rows, 11)
                        wcell1.value = doc1.cell(row=j, column=11).value
                        wcell1 = wc.cell(rows, 12)
                        wcell1.value = doc1.cell(row=j, column=12).value
                        wcell1 = wc.cell(rows, 13)
                        wcell1.value = doc1.cell(row=j, column=13).value
                        wcell1 = wc.cell(rows, 14)
                        wcell1.value = doc1.cell(row=j, column=14).value
                        wcell1 = wc.cell(rows, 15)
                        wcell1.value = doc1.cell(row=j, column=15).value
                        wcell1 = wc.cell(rows, 16)
                        wcell1.value = doc1.cell(row=j, column=16).value
                        wcell1 = wc.cell(rows, 17)
                        wcell1.value = doc1.cell(row=j, column=17).value
                        wcell1 = wc.cell(rows, 18)
                        wcell1.value = doc1.cell(row=j, column=18).value
                        wcell1 = wc.cell(rows, 19)
                        wcell1.value = doc1.cell(row=j, column=19).value
                        wcell1 = wc.cell(rows, 20)
                        wcell1.value = doc1.cell(row=j, column=20).value
                        wcell1 = wc.cell(rows, 21)
                        wcell1.value = doc1.cell(row=j, column=21).value
                        wcell1 = wc.cell(rows, 22)
                        wcell1.value = doc1.cell(row=j, column=22).value
                        wcell1 = wc.cell(rows, 23)
                        wcell1.value = doc1.cell(row=j, column=23).value
                        wcell1 = wc.cell(rows, 24)
                        wcell1.value = doc1.cell(row=j, column=24).value

                        rows = rows + 1
            rows = rows + 2
            p = 0
            for sheet in wz:
                wz.active = sheet
                doc3: Worksheet = wz.active

                for u in range(1, doc3.max_row + 1):


                    if k == doc3.cell(row=u, column=1).value:
                        sheetnames = wt.sheetnames
                        sheetequal = False
                        for l in range(0, len(sheetnames)):
                            if sheetnames[l] == k:
                                sheetequal = True
                        if sheetequal != True:
                            wt.create_sheet(k[:30])
                        wc = wt[k]
                        p = p + 1
                        if p == 1:
                            wcell1 = wc.cell(1, 1)
                            wcell1.value = k
                            wcell1 = wc.cell(rows-1, 5)
                            wcell1.value = "Семестр 1 Заочна"
                            wcell1 = wc.cell(rows, 1)
                            wcell1.value = "Назва навчальних дисциплін"
                            wcell4 = wc.cell(rows, 2)
                            wcell4.value = "К-сть студ"
                            wcell4 = wc.cell(rows, 3)
                            wcell4.value = "Шифр групп"
                            wcell4 = wc.cell(rows, 4)
                            wcell4.value = "К-сть Потоків"
                            wcell4 = wc.cell(rows, 5)
                            wcell4.value = "К-сть Підгруп"
                            wcell2 = wc.cell(rows, 6)
                            wcell2.value = "Чит. лекцій"
                            wcell1 = wc.cell(rows, 7)
                            wcell1.value = "Провед. лабор. занять"
                            wcell1 = wc.cell(rows, 8)
                            wcell1.value = "Провед. практ./ семінар. занять"
                            wcell4 = wc.cell(rows, 9)
                            wcell4.value = "Пров. консульт з нав. дисц. протягом семестру"
                            wcell6 = wc.cell(rows, 10)
                            wcell6.value = "Пров. екзам. консультацій"
                            wcell8 = wc.cell(rows, 11)
                            wcell8.value = "Керівництво і приймання КП/КР"
                            wcell10 = wc.cell(rows, 12)
                            wcell10.value = "Пров. заліку"
                            wcell12 = wc.cell(rows, 13)
                            wcell12.value = "Пров. сем. екз."
                            wcell14 = wc.cell(rows, 14)
                            wcell14.value = "Кер-тво, консульт., реце-ня ДП"
                            wcell4 = wc.cell(rows, 15)
                            wcell4.value = "Пров-ня захисту"
                            wcell4 = wc.cell(rows, 16)
                            wcell4.value = "Кваліф. Іспит"
                            wcell4 = wc.cell(rows, 17)
                            wcell4.value = "Кер-тво НДРС"
                            wcell4 = wc.cell(rows, 18)
                            wcell4.value = "Кер-тво аспірантами, здобувачами"
                            wcell4 = wc.cell(rows, 19)
                            wcell4.value = "Кер-тво практ."
                            wcell1 = wc.cell(rows, 20)
                            wcell1.value = "Інші види -5%"
                            wcell4 = wc.cell(rows, 21)
                            wcell4.value = "Дист. Модуль"
                            wcell1 = wc.cell(rows, 22)
                            wcell1.value = "Додаткові години"
                            wcell1 = wc.cell(rows, 23)
                            wcell1.value = "Всього"
                            rows = rows + 1

                        wcell1 = wc.cell(rows, 1)
                        wcell1.value = doc3.cell(row=1, column=1).value
                        wcell1 = wc.cell(rows, 2)
                        wcell1.value = doc3.cell(row=u, column=2).value
                        wcell1 = wc.cell(rows, 3)
                        wcell1.value = doc3.cell(row=u, column=3).value
                        wcell1 = wc.cell(rows, 4)
                        wcell1.value = doc3.cell(row=u, column=4).value
                        wcell1 = wc.cell(rows, 5)
                        wcell1.value = doc3.cell(row=u, column=5).value
                        wcell1 = wc.cell(rows, 6)
                        wcell1.value = doc3.cell(row=u, column=6).value
                        wcell1 = wc.cell(rows, 7)
                        wcell1.value = doc3.cell(row=u, column=7).value
                        wcell1 = wc.cell(rows, 8)
                        wcell1.value = doc3.cell(row=u, column=8).value
                        wcell1 = wc.cell(rows, 9)
                        wcell1.value = doc3.cell(row=u, column=9).value
                        wcell1 = wc.cell(rows, 10)
                        wcell1.value = doc3.cell(row=u, column=10).value
                        wcell1 = wc.cell(rows, 11)
                        wcell1.value = doc3.cell(row=u, column=11).value
                        wcell1 = wc.cell(rows, 12)
                        wcell1.value = doc3.cell(row=u, column=12).value
                        wcell1 = wc.cell(rows, 13)
                        wcell1.value = doc3.cell(row=u, column=13).value
                        wcell1 = wc.cell(rows, 14)
                        wcell1.value = doc3.cell(row=u, column=14).value
                        wcell1 = wc.cell(rows, 15)
                        wcell1.value = doc3.cell(row=u, column=15).value
                        wcell1 = wc.cell(rows, 16)
                        wcell1.value = doc3.cell(row=u, column=16).value
                        wcell1 = wc.cell(rows, 17)
                        wcell1.value = doc3.cell(row=u, column=17).value
                        wcell1 = wc.cell(rows, 18)
                        wcell1.value = doc3.cell(row=u, column=18).value
                        wcell1 = wc.cell(rows, 19)
                        wcell1.value = doc3.cell(row=u, column=19).value
                        wcell1 = wc.cell(rows, 20)
                        wcell1.value = doc3.cell(row=u, column=20).value
                        wcell1 = wc.cell(rows, 21)
                        wcell1.value = doc3.cell(row=u, column=21).value
                        wcell1 = wc.cell(rows, 22)
                        wcell1.value = doc3.cell(row=u, column=22).value
                        wcell1 = wc.cell(rows, 23)
                        wcell1.value = doc3.cell(row=u, column=23).value
                        wcell1 = wc.cell(rows, 24)
                        wcell1.value = doc3.cell(row=u, column=24).value

                        rows = rows + 1
            rows = rows + 2
            p = 0
            for sheet in wq:
                wq.active = sheet
                doc2: Worksheet = wq.active

                for o in range(1, doc2.max_row + 1):


                    if k == doc2.cell(row=o, column=1).value:

                        sheetnames = wt.sheetnames
                        sheetequal = False
                        for l in range(0, len(sheetnames)):
                            if sheetnames[l] == k:
                                sheetequal = True
                        if sheetequal != True:
                            wt.create_sheet(k[:30])
                        wc = wt[k]
                        p = p + 1
                        if p == 1:
                            wcell1 = wc.cell(1, 1)
                            wcell1.value = k
                            wcell1 = wc.cell(rows-1, 5)
                            wcell1.value = "Семестр 2 Денна"
                            wcell1 = wc.cell(rows, 1)
                            wcell1.value = "Назва навчальних дисциплін"
                            wcell4 = wc.cell(rows, 2)
                            wcell4.value = "К-сть студ"
                            wcell4 = wc.cell(rows, 3)
                            wcell4.value = "Шифр групп"
                            wcell4 = wc.cell(rows, 4)
                            wcell4.value = "К-сть Потоків"
                            wcell4 = wc.cell(rows, 5)
                            wcell4.value = "К-сть Підгруп"
                            wcell2 = wc.cell(rows, 6)
                            wcell2.value = "Чит. лекцій"
                            wcell1 = wc.cell(rows, 7)
                            wcell1.value = "Провед. лабор. занять"
                            wcell1 = wc.cell(rows, 8)
                            wcell1.value = "Провед. практ./ семінар. занять"
                            wcell4 = wc.cell(rows, 9)
                            wcell4.value = "Пров. консульт з нав. дисц. протягом семестру"
                            wcell6 = wc.cell(rows, 10)
                            wcell6.value = "Пров. екзам. консультацій"
                            wcell8 = wc.cell(rows, 11)
                            wcell8.value = "Керівництво і приймання КП/КР"
                            wcell10 = wc.cell(rows, 12)
                            wcell10.value = "Пров. заліку"
                            wcell12 = wc.cell(rows, 13)
                            wcell12.value = "Пров. сем. екз."
                            wcell14 = wc.cell(rows, 14)
                            wcell14.value = "Кер-тво, консульт., реце-ня ДП"
                            wcell4 = wc.cell(rows, 15)
                            wcell4.value = "Пров-ня захисту"
                            wcell4 = wc.cell(rows, 16)
                            wcell4.value = "Кваліф. Іспит"
                            wcell4 = wc.cell(rows, 17)
                            wcell4.value = "Кер-тво НДРС"
                            wcell4 = wc.cell(rows, 18)
                            wcell4.value = "Кер-тво аспірантами, здобувачами"
                            wcell4 = wc.cell(rows, 19)
                            wcell4.value = "Кер-тво практ."
                            wcell1 = wc.cell(rows, 20)
                            wcell1.value = "Інші види -5%"
                            wcell4 = wc.cell(rows, 21)
                            wcell4.value = "Дист. Модуль"
                            wcell1 = wc.cell(rows, 22)
                            wcell1.value = "Додаткові години"
                            wcell1 = wc.cell(rows, 23)
                            wcell1.value = "Всього"
                            rows = rows + 1

                        wcell1 = wc.cell(rows, 1)
                        wcell1.value = doc2.cell(row=1, column=1).value
                        wcell1 = wc.cell(rows, 2)
                        wcell1.value = doc2.cell(row=o, column=2).value
                        wcell1 = wc.cell(rows, 3)
                        wcell1.value = doc2.cell(row=o, column=3).value
                        wcell1 = wc.cell(rows, 4)
                        wcell1.value = doc2.cell(row=o, column=4).value
                        wcell1 = wc.cell(rows, 5)
                        wcell1.value = doc2.cell(row=o, column=5).value
                        wcell1 = wc.cell(rows, 6)
                        wcell1.value = doc2.cell(row=o, column=6).value
                        wcell1 = wc.cell(rows, 7)
                        wcell1.value = doc2.cell(row=o, column=7).value
                        wcell1 = wc.cell(rows, 8)
                        wcell1.value = doc2.cell(row=o, column=8).value
                        wcell1 = wc.cell(rows, 9)
                        wcell1.value = doc2.cell(row=o, column=9).value
                        wcell1 = wc.cell(rows, 10)
                        wcell1.value = doc2.cell(row=o, column=10).value
                        wcell1 = wc.cell(rows, 11)
                        wcell1.value = doc2.cell(row=o, column=11).value
                        wcell1 = wc.cell(rows, 12)
                        wcell1.value = doc2.cell(row=o, column=12).value
                        wcell1 = wc.cell(rows, 13)
                        wcell1.value = doc2.cell(row=o, column=13).value
                        wcell1 = wc.cell(rows, 14)
                        wcell1.value = doc2.cell(row=o, column=14).value
                        wcell1 = wc.cell(rows, 15)
                        wcell1.value = doc2.cell(row=o, column=15).value
                        wcell1 = wc.cell(rows, 16)
                        wcell1.value = doc2.cell(row=o, column=16).value
                        wcell1 = wc.cell(rows, 17)
                        wcell1.value = doc2.cell(row=o, column=17).value
                        wcell1 = wc.cell(rows, 18)
                        wcell1.value = doc2.cell(row=o, column=18).value
                        wcell1 = wc.cell(rows, 19)
                        wcell1.value = doc2.cell(row=o, column=19).value
                        wcell1 = wc.cell(rows, 20)
                        wcell1.value = doc2.cell(row=o, column=20).value
                        wcell1 = wc.cell(rows, 21)
                        wcell1.value = doc2.cell(row=o, column=21).value
                        wcell1 = wc.cell(rows, 22)
                        wcell1.value = doc2.cell(row=o, column=22).value
                        wcell1 = wc.cell(rows, 23)
                        wcell1.value = doc2.cell(row=o, column=23).value
                        wcell1 = wc.cell(rows, 24)
                        wcell1.value = doc2.cell(row=o, column=24).value

                        rows = rows + 1
            rows = rows + 2
            p = 0
            for sheet in wr:
                wr.active = sheet
                doc4: Worksheet = wr.active

                for q in range(1, doc4.max_row + 1):

                    if k == doc4.cell(row=q, column=1).value:
                        sheetnames = wt.sheetnames
                        sheetequal = False
                        for l in range(0, len(sheetnames)):
                            if sheetnames[l] == k:
                                sheetequal = True
                        if sheetequal != True:
                            wt.create_sheet(k[:30])
                        wc = wt[k]
                        p = p + 1
                        if p == 1:
                            wcell1 = wc.cell(1, 1)
                            wcell1.value = k
                            wcell1 = wc.cell(rows-1, 5)
                            wcell1.value = "Семестр 2 Заочна"
                            wcell1 = wc.cell(rows, 1)
                            wcell1.value = "Назва навчальних дисциплін"
                            wcell4 = wc.cell(rows, 2)
                            wcell4.value = "К-сть студ"
                            wcell4 = wc.cell(rows, 3)
                            wcell4.value = "Шифр групп"
                            wcell4 = wc.cell(rows, 4)
                            wcell4.value = "К-сть Потоків"
                            wcell4 = wc.cell(rows, 5)
                            wcell4.value = "К-сть Підгруп"
                            wcell2 = wc.cell(rows, 6)
                            wcell2.value = "Чит. лекцій"
                            wcell1 = wc.cell(rows, 7)
                            wcell1.value = "Провед. лабор. занять"
                            wcell1 = wc.cell(rows, 8)
                            wcell1.value = "Провед. практ./ семінар. занять"
                            wcell4 = wc.cell(rows, 9)
                            wcell4.value = "Пров. консульт з нав. дисц. протягом семестру"
                            wcell6 = wc.cell(rows, 10)
                            wcell6.value = "Пров. екзам. консультацій"
                            wcell8 = wc.cell(rows, 11)
                            wcell8.value = "Керівництво і приймання КП/КР"
                            wcell10 = wc.cell(rows, 12)
                            wcell10.value = "Пров. заліку"
                            wcell12 = wc.cell(rows, 13)
                            wcell12.value = "Пров. сем. екз."
                            wcell14 = wc.cell(rows, 14)
                            wcell14.value = "Кер-тво, консульт., реце-ня ДП"
                            wcell4 = wc.cell(rows, 15)
                            wcell4.value = "Пров-ня захисту"
                            wcell4 = wc.cell(rows, 16)
                            wcell4.value = "Кваліф. Іспит"
                            wcell4 = wc.cell(rows, 17)
                            wcell4.value = "Кер-тво НДРС"
                            wcell4 = wc.cell(rows, 18)
                            wcell4.value = "Кер-тво аспірантами, здобувачами"
                            wcell4 = wc.cell(rows, 19)
                            wcell4.value = "Кер-тво практ."
                            wcell1 = wc.cell(rows, 20)
                            wcell1.value = "Інші види -5%"
                            wcell4 = wc.cell(rows, 21)
                            wcell4.value = "Дист. Модуль"
                            wcell1 = wc.cell(rows, 22)
                            wcell1.value = "Додаткові години"
                            wcell1 = wc.cell(rows, 23)
                            wcell1.value = "Всього"
                            rows = rows + 1
                        wcell1 = wc.cell(rows, 1)
                        wcell1.value = doc4.cell(row=1, column=1).value
                        wcell1 = wc.cell(rows, 2)
                        wcell1.value = doc4.cell(row=q, column=2).value
                        wcell1 = wc.cell(rows, 3)
                        wcell1.value = doc4.cell(row=q, column=3).value
                        wcell1 = wc.cell(rows, 4)
                        wcell1.value = doc4.cell(row=q, column=4).value
                        wcell1 = wc.cell(rows, 5)
                        wcell1.value = doc4.cell(row=q, column=5).value
                        wcell1 = wc.cell(rows, 6)
                        wcell1.value = doc4.cell(row=q, column=6).value
                        wcell1 = wc.cell(rows, 7)
                        wcell1.value = doc4.cell(row=q, column=7).value
                        wcell1 = wc.cell(rows, 8)
                        wcell1.value = doc4.cell(row=q, column=8).value
                        wcell1 = wc.cell(rows, 9)
                        wcell1.value = doc4.cell(row=q, column=9).value
                        wcell1 = wc.cell(rows, 10)
                        wcell1.value = doc4.cell(row=q, column=10).value
                        wcell1 = wc.cell(rows, 11)
                        wcell1.value = doc4.cell(row=q, column=11).value
                        wcell1 = wc.cell(rows, 12)
                        wcell1.value = doc4.cell(row=q, column=12).value
                        wcell1 = wc.cell(rows, 13)
                        wcell1.value = doc4.cell(row=q, column=13).value
                        wcell1 = wc.cell(rows, 14)
                        wcell1.value = doc4.cell(row=q, column=14).value
                        wcell1 = wc.cell(rows, 15)
                        wcell1.value = doc4.cell(row=q, column=15).value
                        wcell1 = wc.cell(rows, 16)
                        wcell1.value = doc4.cell(row=q, column=16).value
                        wcell1 = wc.cell(rows, 17)
                        wcell1.value = doc4.cell(row=q, column=17).value
                        wcell1 = wc.cell(rows, 18)
                        wcell1.value = doc4.cell(row=q, column=18).value
                        wcell1 = wc.cell(rows, 19)
                        wcell1.value = doc4.cell(row=q, column=19).value
                        wcell1 = wc.cell(rows, 20)
                        wcell1.value = doc4.cell(row=q, column=20).value
                        wcell1 = wc.cell(rows, 21)
                        wcell1.value = doc4.cell(row=q, column=21).value
                        wcell1 = wc.cell(rows, 22)
                        wcell1.value = doc4.cell(row=q, column=22).value
                        wcell1 = wc.cell(rows, 23)
                        wcell1.value = doc4.cell(row=q, column=23).value
                        wcell1 = wc.cell(rows, 24)
                        wcell1.value = doc4.cell(row=q, column=24).value

                        rows = rows + 1


        wt.save("zvit2.xlsx")
        wt.close()
        wb.close()
        wd.close()



