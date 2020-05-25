from openpyxl import *


class DennaSem2:
    def __init__(self):
        wb = load_workbook("resources\\ІТтаКБ. Сем II. Форма навчання  денна.xlsx")
        sheet = wb.active

        c = 0
        self.a = []
        i = 0
        j = 0

        for j in range(6, 35):
            self.a.append(c)
            c = 0
            for i in range(18, 64):
                if sheet.cell(row=i, column=j).value is not None \
                        and isinstance(sheet.cell(row=i, column=j).value, str) != True:
                    if i != 55 and i != 56 and i != 57 and i != 58 and i != 59 and i != 60 and i != 61 and i != 63:
                        c = c + sheet.cell(row=i, column=j).value

        print(self.a)
        wb.close()
