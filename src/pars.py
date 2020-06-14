from openpyxl.styles import *
from openpyxl import *
from openpyxl.worksheet.worksheet import Worksheet


class pars:
    def __init__(self, rnp_file, forma_1sem, forma_2sem):

        wb = load_workbook(rnp_file)
        wd = load_workbook(forma_1sem)
        ws = load_workbook(forma_2sem)
        sem2: Worksheet = ws.active
        RNP: Worksheet = wb.active
        sem1: Worksheet = wd.active


        for i in range(1, RNP.max_row):
            # ПРОВЕРКА ПО ПЕРВОМУ ФАЙЛУ
            for j in range(1, sem1.max_row):
                if sem1.cell(row=j, column=2).value == RNP.cell(row=i, column=2).value and sem1.cell(row=j, column=2).value != None:
                    if RNP.cell(row=i, column=32).value == sem1.cell(row=j, column=13).value or RNP.cell(row=i, column=34).value == sem1.cell(row=j, column=13).value:
                        print(i, RNP.cell(row=i, column=2).value)
                        print(j, sem1.cell(row=j, column=2).value)
                        print("KP,KR")
                    elif RNP.cell(row=i, column=32).value == None and sem1.cell(row=j, column=13).value == "--" or RNP.cell(row=i, column=34).value == None and sem1.cell(row=j, column=13).value == "--":
                        print(i, RNP.cell(row=i, column=2).value, RNP.cell(row=i, column=24).value, RNP.cell(row=i, column=26).value, RNP.cell(row=i, column=28).value)
                        print(j, sem1.cell(row=j, column=2).value, sem1.cell(row=j, column=10).value, sem1.cell(row=j, column=11).value, sem1.cell(row=j, column=12).value)

                        if RNP.cell(row=i, column=24).value != sem1.cell(row=j, column=10).value:
                            print('RNP color 24')
                            sem1.cell(row=j, column=10).fill = PatternFill(fill_type='solid', start_color='ff8327', end_color='ff8327')  # Данный код позволяет делать оформление цветом ячейки

                        if RNP.cell(row=i, column=26).value != sem1.cell(row=j, column=11).value:
                            print('RNP color 26')
                            sem1.cell(row=j, column=11).fill = PatternFill(fill_type='solid', start_color='ff8327', end_color='ff8327')  # Данный код позволяет делать оформление цветом ячейки

                        if RNP.cell(row=i, column=28).value != sem1.cell(row=j, column=12).value:
                            print('RNP color 28')
                            sem1.cell(row=j, column=12).fill = PatternFill(fill_type='solid', start_color='ff8327', end_color='ff8327')  # Данный код позволяет делать оформление цветом ячейки
            # ПРОВЕРКА ПО ВТОРОМУ ФАЙЛУ
            for l in range(1, sem2.max_row):
                if sem2.cell(row=l, column=2).value == RNP.cell(row=i, column=2).value and sem2.cell(row=l, column=2).value != None:

                    if RNP.cell(row=i, column=52).value == sem2.cell(row=l, column=13).value or RNP.cell(row=i, column=54).value == sem2.cell(row=l, column=13).value:
                            print(i, RNP.cell(row=i, column=2).value)
                            print(l, sem2.cell(row=l, column=2).value)
                            print("KP,KR")
                    elif RNP.cell(row=i, column=52).value == None and sem2.cell(row=l, column=13).value == "--" or RNP.cell(row=i, column=54).value == None and sem2.cell(row=l, column=13).value == "--":
                        print(i, RNP.cell(row=i, column=2).value, RNP.cell(row=i, column=44).value, RNP.cell(row=i, column=46).value, RNP.cell(row=i, column=48).value)
                        print(l, sem2.cell(row=l, column=2).value, sem2.cell(row=l, column=10).value, sem2.cell(row=l, column=11).value, sem2.cell(row=l, column=12).value)

                        if RNP.cell(row=i, column=44).value != sem2.cell(row=l, column=10).value:
                             sem2.cell(row=l, column=10).fill = PatternFill(fill_type='solid', start_color='ff8327', end_color='ff8327')  # Данный код позволяет делать оформление цветом ячейки

                        if RNP.cell(row=i, column=46).value != sem2.cell(row=l, column=11).value:
                            sem2.cell(row=l, column=11).fill = PatternFill(fill_type='solid', start_color='ff8327', end_color='ff8327')  # Данный код позволяет делать оформление цветом ячейки

                        if RNP.cell(row=i, column=48).value != sem2.cell(row=l, column=12).value:
                            sem2.cell(row=l, column=12).fill = PatternFill(fill_type='solid', start_color='ff8327', end_color='ff8327')  # Данный код позволяет делать оформление цветом ячейки



        wb.save(rnp_file)
        wd.save(forma_1sem)
        ws.save(forma_2sem)
        wb.close()
        wd.close()
        ws.close()
