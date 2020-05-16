from openpyxl import *


class ZaoсhnaSem1:
    def __init__(self):
        wb = load_workbook("resources\\ІТтаКБ. Сем I. Форма навчання  заочна.xlsx")
        print("ZAOCHNA")
        sheet = wb.active
        print(wb.get_sheet_names())
        b = []
        r = 0
        for j in range(6, 35):
            b.append(r)
            r = 0
            for i in range(18, 51):
                print(i, sheet.cell(row=i, column=j).value)
                if isinstance(sheet.cell(row=i, column=j).value, str) != True and sheet.cell(row=i, column=j).value is not None:
                    r = r + sheet.cell(row=i, column=j).value

        print(b)
        wb.close()
