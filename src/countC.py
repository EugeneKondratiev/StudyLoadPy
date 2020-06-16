from openpyxl import *
from openpyxl.worksheet.worksheet import Worksheet

import numpy as np


class CountC:
    def __init__(self, file_name_1sem_den, file_name_1sem_zaoch, file_name_2sem_den, file_name_2sem_zaoch, subject):
        wd = load_workbook(file_name_1sem_den)
        ws = load_workbook(file_name_1sem_zaoch)
        wq = load_workbook(file_name_2sem_den)
        ww = load_workbook(file_name_2sem_zaoch)
        denna2: Worksheet = wq.active
        zaochna2: Worksheet = ww.active
        denna1: Worksheet = wd.active
        zaochna1: Worksheet = ws.active
        predmet = subject
        lektor_d1 = []
        lektor_d2 = []
        lektor_z1 = []
        lektor_z2 = []
        ass_den1 = []
        ass_den2 = []
        ass_zaoch1 = []
        ass_zaoch2 = []
        den = {}
        zaoch = {}
        c = []
        podgrup_d1 = 0
        podgrup_d2 = 0
        podgrup_z1 = 0
        podgrup_z2 = 0
        chasy_den1 = []
        chasy_zaoch1 = []
        chasy_den2 = []
        chasy_zaoch2 = []


        for i in range(18, denna1.max_row):
            if denna1.cell(row=i, column=2).value != None and denna1.cell(row=i, column=2).value != 0:
                k = denna1.cell(row=i - 1, column=2).value

                dict1_cond = {temp: v for (temp, v) in den.items() if temp == k if temp != None}
                if len(dict1_cond.items()) != 0 and t != "":
                    nympyarray1 = np.array(list(dict1_cond.values()))
                    nympyarray2 = np.array(c)
                    nympyarray3 = np.array(nympyarray1 + nympyarray2)
                    den[t] = nympyarray3.tolist()
                else:
                    den[k] = c

                c = []
                t = k
                for j in range(16, denna1.max_column + 1):
                    if denna1.cell(row=i, column=j).value is not None and isinstance(denna1.cell(row=i, column=j).value,
                                                                                     str) != True:
                        c.append(denna1.cell(row=i, column=j).value)
        c = []
        t = ""
        y = 0
        for i in range(18, zaochna1.max_row):
            if zaochna1.cell(row=i, column=2).value != None and zaochna1.cell(row=i, column=2).value != 0:
                k = zaochna1.cell(row=i - 1, column=2).value
                dict2_cond = {temp: v for (temp, v) in zaoch.items() if temp == k if temp != None}
                if len(dict2_cond.items()) != 0 and t != "":
                    nympyarray1 = np.array(list(dict2_cond.values()))
                    nympyarray2 = np.array(c)
                    nympyarray3 = np.array(nympyarray1 + nympyarray2)
                    zaoch[t] = nympyarray3.tolist()
                else:
                    zaoch[k] = c
                t = k
                c = []
                for j in range(16, zaochna1.max_column + 1):
                    if zaochna1.cell(row=i, column=j).value is not None and isinstance(
                            zaochna1.cell(row=i, column=j).value, str) != True:
                        c.append(zaochna1.cell(row=i, column=j).value)
        del den[None]
        del zaoch[None]

        den2 = {}
        zaoch2 = {}
        c = []
        k = ""
        for i in range(18, denna2.max_row):
            if denna2.cell(row=i, column=2).value != None and denna2.cell(row=i, column=2).value != 0:
                k = denna2.cell(row=i - 1, column=2).value

                dict3_cond = {temp: v for (temp, v) in den2.items() if temp == k if temp != None}
                if len(dict3_cond.items()) != 0 and t != "":
                    nympyarray1 = np.array(list(dict3_cond.values()))
                    nympyarray2 = np.array(c)
                    nympyarray3 = np.array(nympyarray1 + nympyarray2)
                    den2[t] = nympyarray3.tolist()
                else:
                    den2[k] = c

                c = []
                t = k
                for j in range(16, denna2.max_column + 1):
                    if denna2.cell(row=i, column=j).value is not None and isinstance(denna2.cell(row=i, column=j).value,
                                                                                     str) != True:
                        c.append(denna2.cell(row=i, column=j).value)
        c = []
        t = ""
        y = 0
        for i in range(18, zaochna2.max_row):
            if zaochna2.cell(row=i, column=2).value != None and zaochna2.cell(row=i, column=2).value != 0:
                k = zaochna2.cell(row=i - 1, column=2).value
                dict4_cond = {temp: v for (temp, v) in zaoch2.items() if temp == k if temp != None}
                if len(dict4_cond.items()) != 0 and t != "":
                    nympyarray1 = np.array(list(dict4_cond.values()))
                    nympyarray2 = np.array(c)
                    nympyarray3 = np.array(nympyarray1 + nympyarray2)
                    zaoch2[t] = nympyarray3.tolist()
                else:
                    zaoch2[k] = c
                t = k
                c = []
                for j in range(16, zaochna2.max_column + 1):
                    if zaochna2.cell(row=i, column=j).value is not None and isinstance(
                            zaochna2.cell(row=i, column=j).value, str) != True:
                        c.append(zaochna2.cell(row=i, column=j).value)

        del den2[None]
        del zaoch2[None]

        for i in range(1, denna1.max_row):
            if predmet == denna1.cell(row=i, column=2).value:
                podgrup_d1 = denna1.cell(row=i, column=7).value

        for i in range(1, denna2.max_row):
            if predmet == denna2.cell(row=i, column=2).value:
                podgrup_d2 = denna2.cell(row=i, column=7).value

        for i in range(1, zaochna1.max_row):
            if predmet == zaochna1.cell(row=i, column=2).value:
                podgrup_z1 = zaochna1.cell(row=i, column=7).value

        for i in range(1, zaochna2.max_row):
            if predmet == zaochna2.cell(row=i, column=2).value:
                podgrup_z2 = zaochna2.cell(row=i, column=7).value

        for (k, c) in den.items():
            if predmet == k:
                chasy_den1 = c

        k = ""
        for (y, c) in zaoch.items():
            if predmet == y:
                chasy_zaoch1 = c

        for (k, c) in den2.items():
            if predmet == k:
                chasy_den2 = c

        k = ""
        for (y, c) in zaoch2.items():
            if predmet == y:
                chasy_zaoch2 = c

        for i in range(0, len(chasy_den1)):
            if i == 0 or i == 3 or i == 4 or i == 5 or i == 6 or i == 7 or i == 8 or i == 9 or i == 10 or i == 11 or i == 12 or i == 15 or i == 16:
                lektor_d1.append(chasy_den1[i])
            elif i != 17:
                ass_den1.append(chasy_den1[i])

        for i in range(0, len(chasy_den2)):
            if i == 0 or i == 3 or i == 4 or i == 5 or i == 6 or i == 7 or i == 8 or i == 9 or i == 10 or i == 11 or i == 12 or i == 15 or i == 16:
                lektor_d2.append(chasy_den2[i])
            elif i != 17:
                ass_den2.append(chasy_den2[i])

        for i in range(0, len(chasy_zaoch1)):
            if i == 0 or i == 3 or i == 4 or i == 5 or i == 6 or i == 7 or i == 8 or i == 9 or i == 10 or i == 11 or i == 12 or i == 15 or i == 16:
                lektor_z1.append(chasy_zaoch1[i])
            elif i != 17:
                ass_zaoch1.append(chasy_zaoch1[i])

        for i in range(0, len(chasy_zaoch2)):
            if i == 0 or i == 3 or i == 4 or i == 5 or i == 6 or i == 7 or i == 8 or i == 9 or i == 10 or i == 11 or i == 12 or i == 15 or i == 16:
                lektor_z2.append(chasy_zaoch2[i])
            elif i != 17:
                ass_zaoch2.append(chasy_zaoch2[i])

        self.podgrup_d1sem = podgrup_d1
        self.podgrup_d2sem = podgrup_d2
        self.podgrup_z1sem = podgrup_z1
        self.podgrup_z2sem = podgrup_z2
        self.sem1_den_ass = np.sum(ass_den1)
        self.sem2_den_ass = np.sum(ass_den2)
        self.sem1_zaoch_ass = np.sum(ass_zaoch1)
        self.sem2_zaoch_ass = np.sum(ass_zaoch2)
        self.sem1_den_lektor = np.sum(lektor_d1)
        self.sem2_den_lektor = np.sum(lektor_d2)
        self.sem1_zaoch_lektor = np.sum(lektor_z1)
        self.sem2_zaoch_lektor = np.sum(lektor_z2)

